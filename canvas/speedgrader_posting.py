"""
Standardized Canvas SpeedGrader prepare/post/verify workflow.

Usage examples:
  python3 -m canvas.speedgrader_posting prepare --section T5 --excel /path/to/final.xlsx
  python3 -m canvas.speedgrader_posting dry-run --section T5
  python3 -m canvas.speedgrader_posting post --section T5
  python3 -m canvas.speedgrader_posting verify --section T5
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
from dataclasses import asdict
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import Page, sync_playwright

from canvas.posting_config import (
    DEFAULT_CANVAS_BASE,
    DEFAULT_PAYLOAD_PATH,
    DEFAULT_STORAGE_STATE,
    SECTION_CONFIGS,
    SectionConfig,
)

HIDE_OVERLAYS_JS = """
document.querySelectorAll('.eesy, .eesy-tab2-container, [id*="eesy"], [class*="eesy"]')
  .forEach(el => el.style.setProperty('display', 'none', 'important'));
"""


def load_section_config(section: str) -> SectionConfig:
    key = section.upper()
    if key not in SECTION_CONFIGS:
        raise SystemExit(f"Unknown section '{section}'. Add it to canvas/posting_config.py first.")
    return SECTION_CONFIGS[key]


def normalize_name(name: str) -> str:
    name = name.lower().strip()
    name = name.replace("_", " ").replace(".", " ")
    if "," in name:
        last, first = [p.strip() for p in name.split(",", 1)]
        name = f"{first} {last}"
    return re.sub(r"[^a-z0-9 ]+", "", " ".join(name.split()))


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, normalize_name(a), normalize_name(b)).ratio()


def extract_q_numbers(text: str) -> list[str]:
    found = re.findall(r"Q\d+", text or "", flags=re.IGNORECASE)
    seen = set()
    result = []
    for item in found:
        q = item.upper()
        if q not in seen:
            seen.add(q)
            result.append(q)
    return result


def build_clean_comment(canvas_comment: str, questions_missed: str, free_response_questions: tuple[str, ...]) -> str:
    fr_set = {q.upper() for q in free_response_questions}
    preferred = extract_q_numbers(canvas_comment)
    if not preferred:
        preferred = extract_q_numbers(questions_missed)
    filtered = [q for q in preferred if q not in fr_set]
    return f"Missed: {', '.join(filtered)}" if filtered else ""


def read_final_workbook(excel_path: Path, free_response_questions: tuple[str, ...]) -> list[dict]:
    wb = load_workbook(excel_path, read_only=True)
    ws = wb["Grades"]

    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {header: idx for idx, header in enumerate(headers) if header}

    name_col = header_map.get("Name", header_map.get("Student"))
    score_col = header_map.get("Score", header_map.get("Raw Score"))
    total_col = header_map.get("/32", header_map.get("Total Points"))
    missed_col = header_map.get("Questions Missed")
    canvas_comment_col = header_map.get("Canvas Comment")
    flags_col = header_map.get("Flags")
    percent_col = header_map.get("%")
    confidence_col = header_map.get("Confidence")

    if name_col is None or score_col is None:
        raise ValueError(f"Could not find Name/Score columns in {excel_path}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[name_col] in (None, ""):
            continue

        try:
            score = float(row[score_col])
        except (TypeError, ValueError):
            continue

        total = None
        if total_col is not None and total_col < len(row) and row[total_col] not in (None, ""):
            total = float(row[total_col])

        questions_missed = str(row[missed_col]).strip() if missed_col is not None and missed_col < len(row) and row[missed_col] else ""
        canvas_comment = str(row[canvas_comment_col]).strip() if canvas_comment_col is not None and canvas_comment_col < len(row) and row[canvas_comment_col] else ""
        flags = str(row[flags_col]).strip() if flags_col is not None and flags_col < len(row) and row[flags_col] else ""
        percent = row[percent_col] if percent_col is not None and percent_col < len(row) else None
        confidence = str(row[confidence_col]).strip() if confidence_col is not None and confidence_col < len(row) and row[confidence_col] else ""

        rows.append({
            "name": str(row[name_col]).strip(),
            "score": score,
            "max": total,
            "percent": percent,
            "missed": questions_missed,
            "flags": flags,
            "canvas_comment": build_clean_comment(canvas_comment, questions_missed, free_response_questions),
            "confidence": confidence,
        })

    wb.close()
    return rows


def load_payload(payload_path: Path, section: str) -> dict:
    if not payload_path.exists():
        raise SystemExit(f"Payload file not found: {payload_path}")
    data = json.loads(payload_path.read_text())
    if section not in data:
        raise SystemExit(f"Section '{section}' not found in payload: {payload_path}")
    return data


def save_payload(payload_path: Path, payload: dict) -> None:
    payload_path.write_text(json.dumps(payload, indent=2))


def wait_for(page: Page, selector: str, timeout: int = 30000):
    return page.wait_for_selector(selector, timeout=timeout)


def hide_overlays(page: Page) -> None:
    try:
        page.evaluate(HIDE_OVERLAYS_JS)
    except Exception:
        pass


def open_speedgrader(page: Page, cfg: SectionConfig, student_id: str | None = None) -> None:
    url = (
        f"{DEFAULT_CANVAS_BASE}/courses/{cfg.course_id}/gradebook/speed_grader"
        f"?assignment_id={cfg.assignment_id}"
    )
    if student_id:
        url += f"&student_id={student_id}"
    page.goto(url)
    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass
    wait_for(page, '[data-testid="student-select-trigger"]')
    hide_overlays(page)
    time.sleep(1.0)


def ensure_assignment_points(page: Page, cfg: SectionConfig) -> str:
    page.goto(f"{DEFAULT_CANVAS_BASE}/courses/{cfg.course_id}/assignments/{cfg.assignment_id}/edit")
    page.wait_for_selector('#assignment_points_possible, input[name="points_possible"]', state="attached", timeout=30000)
    hide_overlays(page)
    field = page.locator('#assignment_points_possible, input[name="points_possible"]').first
    current = (field.input_value() or "").strip()
    expected = str(int(cfg.exam_max) if float(cfg.exam_max).is_integer() else cfg.exam_max)
    if current != expected:
        field.click()
        field.press("Meta+A")
        field.type(expected)
        save_button = page.locator('button:has-text("Save")').first
        save_button.click()
        try:
            page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass
        time.sleep(2.0)
        page.goto(f"{DEFAULT_CANVAS_BASE}/courses/{cfg.course_id}/assignments/{cfg.assignment_id}/edit")
        page.wait_for_selector('#assignment_points_possible, input[name="points_possible"]', state="attached", timeout=30000)
        current = (page.locator('#assignment_points_possible, input[name="points_possible"]').first.input_value() or "").strip()
        if current != expected:
            raise RuntimeError(f"Assignment points are '{current}', expected '{expected}'")
    return current


def get_roster_names(page: Page, cfg: SectionConfig) -> list[str]:
    open_speedgrader(page, cfg)
    trigger = page.locator('[data-testid="student-select-trigger"]')
    trigger.click()
    wait_for(page, '[data-testid^="student-option-"]')
    options = page.locator('[data-testid^="student-option-"]')
    names = []
    for idx in range(options.count()):
        text = options.nth(idx).inner_text().strip()
        if text:
            names.append(text)
    page.keyboard.press("Escape")
    return names


def build_matches(students: list[dict], roster_names: list[str], name_overrides: dict[str, str] | None = None) -> list[dict]:
    used = set()
    matches = []
    for student in students:
        override_name = None
        if name_overrides and student["name"] in name_overrides:
            override_name = name_overrides[student["name"]]
        best_name = None
        best_score = 0.0
        matched = dict(student)
        if override_name:
            matched["canvas_name"] = override_name
            matched["match_score"] = 1.0
        else:
            for roster_name in roster_names:
                if roster_name in used:
                    continue
                score = similarity(student["name"], roster_name)
                if score > best_score:
                    best_name = roster_name
                    best_score = score
            matched["canvas_name"] = best_name if best_score >= 0.60 else None
            matched["match_score"] = round(best_score, 2)
        if matched["canvas_name"]:
            used.add(matched["canvas_name"])
        matches.append(matched)
    return matches


def select_student_by_name(page: Page, canvas_name: str) -> str:
    hide_overlays(page)
    page.locator('[data-testid="student-select-trigger"]').click()
    wait_for(page, '[data-testid^="student-option-"]')
    options = page.locator('[data-testid^="student-option-"]')
    clicked = False
    for idx in range(options.count()):
        text = options.nth(idx).inner_text().strip()
        if text == canvas_name:
            options.nth(idx).click()
            clicked = True
            break
    if not clicked:
        raise RuntimeError(f"Could not find exact roster option for '{canvas_name}'")
    try:
        page.wait_for_load_state("networkidle", timeout=10000)
    except Exception:
        pass
    time.sleep(1.0)
    m = re.search(r"[?&]student_id=(\d+)", page.url)
    if not m:
        raise RuntimeError(f"Could not determine Canvas student id after selecting '{canvas_name}'")
    hide_overlays(page)
    return m.group(1)


def current_grade(page: Page) -> str:
    hide_overlays(page)
    return (page.locator('[data-testid="grade-input"]').first.input_value() or "").strip()


def read_page_text(page: Page, expected_comment: str = "") -> str:
    text = ""
    for _ in range(4):
        text = page.locator("body").inner_text(timeout=5000)
        if (
            (expected_comment and expected_comment in text)
            or "Delete comment:" in text
            or "Draft" in text
        ):
            break
        time.sleep(1.0)
    return text


def set_grade(page: Page, score: float) -> None:
    grade = page.locator('[data-testid="grade-input"]').first
    score_text = str(int(score) if float(score).is_integer() else score)
    grade.click()
    grade.press("Meta+A")
    grade.type(score_text)
    grade.press("Tab")
    time.sleep(1.2)


def count_delete_buttons(page: Page) -> int:
    hide_overlays(page)
    return page.locator('[data-testid$="-delete-button"]').count()


def count_visible_comments(page_text: str) -> int:
    return page_text.count("Delete comment:")


def clear_existing_comments(page: Page) -> int:
    deleted = 0
    for _ in range(30):
        hide_overlays(page)
        if page.locator('[data-testid$="-delete-button"]').count() == 0:
            break
        page.evaluate(
            """
            () => {
              const btn = document.querySelector('[data-testid$="-delete-button"]');
              if (btn) btn.click();
            }
            """
        )
        time.sleep(1.5)
        deleted += 1
    return deleted


def clear_rce_body(page: Page) -> None:
    for frame in page.frames:
        if "_ifr" in frame.name and "rce" in frame.name:
            body = frame.locator("body")
            body.click()
            page.keyboard.press("Meta+A")
            page.keyboard.press("Backspace")
            return
    area = page.locator("textarea").first
    if area.count():
        area.click()
        area.press("Meta+A")
        area.press("Backspace")


def post_comment(page: Page, comment: str) -> None:
    if not comment:
        raise RuntimeError("Refusing to post an empty comment")

    hide_overlays(page)
    typed = False
    for frame in page.frames:
        if "_ifr" in frame.name and "rce" in frame.name:
            body = frame.locator("body")
            body.click()
            time.sleep(0.2)
            page.keyboard.press("Meta+A")
            page.keyboard.press("Backspace")
            page.keyboard.type(comment, delay=15)
            typed = True
            break

    if not typed:
        area = page.locator("textarea").first
        if area.count() == 0:
            raise RuntimeError("Could not find Canvas comment editor")
        area.click()
        time.sleep(0.2)
        area.press("Meta+A")
        area.type(comment, delay=15)

    hide_overlays(page)
    result = page.evaluate(
        """
        () => {
          const buttons = Array.from(document.querySelectorAll('button'));
          const visible = buttons.filter(btn => {
            const r = btn.getBoundingClientRect();
            return r.width > 0 && r.height > 0 && !btn.disabled;
          });

          const textMatch = visible.find(btn => btn.innerText.trim() === 'Submit Comment');
          if (textMatch) { textMatch.click(); return 'Submit Comment'; }

          const primary = visible.find(btn => btn.getAttribute('data-testid') === 'submit-comment-button');
          if (primary) { primary.click(); return 'submit-comment-button'; }

          const fallback = visible.find(btn => btn.id === 'comment_submit_button');
          if (fallback) { fallback.click(); return 'comment_submit_button'; }

          return null;
        }
        """
    )
    if not result:
        raise RuntimeError("Could not find an enabled comment submit button")
    time.sleep(2.0)

    page_text = page.locator("body").inner_text(timeout=5000)
    if "Draft" in page_text:
        draft_submit = page.locator('button[data-testid$="-submit-button"]').first
        if draft_submit.count():
            draft_submit.click(force=True)
            time.sleep(2.0)


def verify_student(page: Page, expected_score: float, expected_comment: str) -> tuple[bool, str]:
    score_text = str(int(expected_score) if float(expected_score).is_integer() else expected_score)
    grade_ok = current_grade(page) == score_text
    page_text = read_page_text(page, expected_comment)
    comment_count = count_visible_comments(page_text)
    comment_ok = expected_comment in page_text
    draft_present = "Draft" in page_text
    problems = []
    if not grade_ok:
        problems.append(f"grade mismatch (saw {current_grade(page)!r}, expected {score_text!r})")
    if comment_count != 1:
        problems.append(f"expected 1 comment, found {comment_count}")
    if not comment_ok:
        problems.append("posted comment text not visible on page")
    if draft_present:
        problems.append("comment still in draft state")
    return (not problems, "; ".join(problems) if problems else "ok")


def open_context(headless: bool, storage_state: Path):
    if not storage_state.exists():
        raise SystemExit(
            f"Canvas auth state not found: {storage_state}\n"
            "Create it first by logging in with a headed Playwright browser and saving storage state."
        )
    pw = sync_playwright().start()
    browser = pw.chromium.launch(headless=headless, slow_mo=0 if headless else 50)
    context = browser.new_context(
        viewport={"width": 1440, "height": 900},
        storage_state=str(storage_state),
    )
    page = context.new_page()
    page.on("dialog", lambda dialog: dialog.accept())
    return pw, browser, context, page


def cmd_prepare(args: argparse.Namespace) -> int:
    cfg = load_section_config(args.section)
    payload_path = Path(args.payload)
    payload = json.loads(payload_path.read_text()) if payload_path.exists() else {}
    payload[cfg.section] = read_final_workbook(Path(args.excel), cfg.free_response_questions)
    save_payload(payload_path, payload)
    print(f"Wrote {len(payload[cfg.section])} records to {payload_path} for {cfg.section}")
    return 0


def cmd_dry_run(args: argparse.Namespace) -> int:
    cfg = load_section_config(args.section)
    payload = load_payload(Path(args.payload), cfg.section)
    students = payload[cfg.section]

    pw, browser, context, page = open_context(headless=args.headless, storage_state=Path(args.storage_state))
    try:
        roster = get_roster_names(page, cfg)
        matches = build_matches(students, roster, cfg.name_overrides)
        weak = 0
        for match in matches:
            print(
                f"{match['name']:<26} -> "
                f"{(match.get('canvas_name') or 'UNMATCHED'):<28} "
                f"{match['match_score']:.2f}"
            )
            if not match.get("canvas_name") or match["match_score"] < 0.80:
                weak += 1

        if args.write_back:
            used_ids = 0
            open_speedgrader(page, cfg)
            for match in matches:
                if match.get("canvas_name"):
                    student_id = select_student_by_name(page, match["canvas_name"])
                    match["canvas_id"] = student_id
                    used_ids += 1
            payload[cfg.section] = matches
            save_payload(Path(args.payload), payload)
            print(f"\nSaved resolved Canvas names/ids for {used_ids} students to {args.payload}")

        print(f"\nRoster count: {len(roster)}")
        print(f"Weak or unmatched matches: {weak}")
        return 0 if weak == 0 else 2
    finally:
        context.close()
        browser.close()
        pw.stop()


def cmd_post(args: argparse.Namespace) -> int:
    cfg = load_section_config(args.section)
    payload = load_payload(Path(args.payload), cfg.section)
    students = payload[cfg.section]

    pw, browser, context, page = open_context(headless=args.headless, storage_state=Path(args.storage_state))
    try:
        points = ensure_assignment_points(page, cfg)
        print(f"Assignment points confirmed at {points}")

        failures = []
        for idx, student in enumerate(students, start=1):
            comment = (student.get("canvas_comment") or "").strip()
            if not comment:
                failures.append((student["name"], "empty canvas_comment"))
                continue

            canvas_id = str(student.get("canvas_id") or "").strip()
            canvas_name = (student.get("canvas_name") or "").strip()

            if canvas_id:
                open_speedgrader(page, cfg, canvas_id)
            else:
                open_speedgrader(page, cfg)
                if not canvas_name:
                    failures.append((student["name"], "missing canvas_id/canvas_name"))
                    continue
                canvas_id = select_student_by_name(page, canvas_name)
                student["canvas_id"] = canvas_id

            print(f"[{idx}/{len(students)}] {student['name']} -> {canvas_name or canvas_id}")
            set_grade(page, float(student["score"]))
            deleted = clear_existing_comments(page)
            post_comment(page, comment)
            ok, detail = verify_student(page, float(student["score"]), comment)
            print(f"  deleted={deleted} verify={detail}")
            if not ok:
                failures.append((student["name"], detail))

        save_payload(Path(args.payload), payload)
        if failures:
            print("\nFailures:")
            for name, detail in failures:
                print(f"  {name}: {detail}")
            return 1
        print("\nPosting complete with no verification failures.")
        return 0
    finally:
        context.close()
        browser.close()
        pw.stop()


def cmd_verify(args: argparse.Namespace) -> int:
    cfg = load_section_config(args.section)
    payload = load_payload(Path(args.payload), cfg.section)
    students = payload[cfg.section]

    pw, browser, context, page = open_context(headless=args.headless, storage_state=Path(args.storage_state))
    try:
        failures = []
        for idx, student in enumerate(students, start=1):
            canvas_id = str(student.get("canvas_id") or "").strip()
            if not canvas_id:
                failures.append((student["name"], "missing canvas_id"))
                continue
            open_speedgrader(page, cfg, canvas_id)
            ok, detail = verify_student(page, float(student["score"]), student.get("canvas_comment", ""))
            print(f"[{idx}/{len(students)}] {student['name']}: {detail}")
            if not ok:
                failures.append((student["name"], detail))

        if failures:
            print("\nVerification failures:")
            for name, detail in failures:
                print(f"  {name}: {detail}")
            return 1
        print("\nVerification complete with no failures.")
        return 0
    finally:
        context.close()
        browser.close()
        pw.stop()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Standardized Canvas SpeedGrader workflow")
    sub = parser.add_subparsers(dest="cmd", required=True)

    prepare = sub.add_parser("prepare", help="Extract clean posting payload from a final Excel workbook")
    prepare.add_argument("--section", required=True)
    prepare.add_argument("--excel", required=True)
    prepare.add_argument("--payload", default=str(DEFAULT_PAYLOAD_PATH))
    prepare.set_defaults(func=cmd_prepare)

    dry_run = sub.add_parser("dry-run", help="Fetch SpeedGrader roster and show name matches")
    dry_run.add_argument("--section", required=True)
    dry_run.add_argument("--payload", default=str(DEFAULT_PAYLOAD_PATH))
    dry_run.add_argument("--storage-state", default=str(DEFAULT_STORAGE_STATE))
    dry_run.add_argument("--write-back", action="store_true", help="Save matched Canvas names and ids into the payload")
    dry_run.add_argument("--headless", action="store_true", help="Run without a visible browser window")
    dry_run.set_defaults(func=cmd_dry_run)

    post = sub.add_parser("post", help="Post grades and one cleaned comment per student")
    post.add_argument("--section", required=True)
    post.add_argument("--payload", default=str(DEFAULT_PAYLOAD_PATH))
    post.add_argument("--storage-state", default=str(DEFAULT_STORAGE_STATE))
    post.add_argument("--headless", action="store_true", help="Run without a visible browser window")
    post.set_defaults(func=cmd_post)

    verify = sub.add_parser("verify", help="Re-open each student in SpeedGrader and verify grade/comment state")
    verify.add_argument("--section", required=True)
    verify.add_argument("--payload", default=str(DEFAULT_PAYLOAD_PATH))
    verify.add_argument("--storage-state", default=str(DEFAULT_STORAGE_STATE))
    verify.add_argument("--headless", action="store_true", help="Run without a visible browser window")
    verify.set_defaults(func=cmd_verify)

    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
