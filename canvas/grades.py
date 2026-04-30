"""
Post grades and comments back to Canvas LMS.
"""

import logging
import re
from difflib import SequenceMatcher
from canvas.client import CanvasClient

logger = logging.getLogger(__name__)


def extract_score(grade_text: str) -> float | None:
    """Extract numeric score from grade output (uses last GRADE: match for total)."""
    matches = re.findall(r"GRADE:\s*([\d.]+)\s*/", grade_text)
    if matches:
        return float(matches[-1])  # last match is the total
    return None


def post_grade(
    client: CanvasClient,
    course_id: int,
    assignment_id: int,
    student_id: int,
    score: float,
    comment: str = "",
) -> dict:
    """
    Post a grade and comment to Canvas for a student.

    Args:
        client: Canvas API client
        course_id: Canvas course ID
        assignment_id: Canvas assignment ID
        student_id: Canvas user ID
        score: Numeric score
        comment: Text comment for the student

    Returns:
        Canvas API response
    """
    data = {
        "submission": {
            "posted_grade": str(score),
        }
    }

    if comment:
        data["comment"] = {
            "text_comment": comment,
        }

    resp = client.put(
        f"courses/{course_id}/assignments/{assignment_id}/submissions/{student_id}",
        data=data,
    )
    return resp.json()


def post_grades_batch(
    client: CanvasClient,
    course_id: int,
    assignment_id: int,
    results: list[dict],
    total_points: float = None,
) -> list[dict]:
    """
    Post grades for multiple students silently (no confirmation prompts).

    Uses the "Canvas Comment" field if available (pre-formatted safe comment from Excel).
    Falls back to extracting missed question numbers from grade_text.

    NEVER posts FR Details or specific wrong answers — only missed Q numbers + general feedback.

    Args:
        client: Canvas API client
        course_id: Canvas course ID
        assignment_id: Canvas assignment ID
        results: List of grade results (must include student_id and grade_text)
        total_points: Total possible points (for display). Auto-detected if None.

    Returns:
        List of posted results
    """
    posted = []
    posted_ids: set[int] = set()

    for result in results:
        if result.get("error"):
            logger.warning("Skipping %s — had grading error", result["student_file"])
            posted.append({
                "student_name": result.get("student_name", result["student_file"]),
                "score": None,
                "posted": False,
                "error": "grading error",
            })
            continue

        # Prefer structured result score, fall back to regex
        sr = result.get("structured_result")
        if sr and hasattr(sr, 'total_score') and sr.total_score is not None:
            score = sr.total_score
            result["_score_source"] = "structured"
        else:
            score = extract_score(result["grade_text"])
            if score is not None:
                result["_score_source"] = "regex_fallback"

        if score is None:
            logger.warning("Skipping %s — couldn't extract score", result["student_file"])
            continue

        # Bounds check: reject impossible scores before posting
        max_score = total_points
        if max_score is None and sr and hasattr(sr, 'total_possible'):
            max_score = sr.total_possible
        if max_score and score > max_score:
            logger.warning("%s score %s > max %s — clamping to %s", result["student_file"], score, max_score, max_score)
            score = max_score
        if score < 0:
            logger.warning("%s negative score %s — clamping to 0", result["student_file"], score)
            score = 0

        student_id = result.get("student_id")
        if not student_id:
            logger.warning("Skipping %s — no student_id", result["student_file"])
            continue

        # Duplicate student_id detection — prevent posting twice to same student
        if student_id in posted_ids:
            logger.warning("DUPLICATE student_id %s for %s — skipping to prevent overwrite", student_id, result["student_file"])
            posted.append({"student_name": result.get("student_name", result["student_file"]),
                          "score": score, "posted": False, "error": "duplicate student_id"})
            continue
        posted_ids.add(student_id)

        student_name = result.get("student_name", result["student_file"])
        display_total = f"/{int(total_points)}" if total_points else ""

        try:
            # Prefer pre-formatted Canvas Comment (from Excel review)
            comment = result.get("canvas_comment", "")
            if not comment:
                comment = build_public_comment(result)
            post_grade(client, course_id, assignment_id, student_id, score, comment)
            logger.info("%s: %s%s -> Submitted", student_name, score, display_total)
            posted.append({"student_name": student_name, "score": score, "posted": True})
        except Exception as e:
            logger.error("%s: %s%s -> ERROR: %s", student_name, score, display_total, e)
            posted.append({"student_name": student_name, "score": score, "posted": False, "error": str(e)})

    return posted


def build_public_comment(result: dict) -> str:
    """
    Build a safe public comment for Canvas from a grading result.

    Includes: missed question numbers + general feedback comments.
    NEVER includes: FR Details, specific wrong answers, rubric reasoning, answer content.
    """
    parts = []

    grade_text = result.get("grade_text", "")

    # Missed questions — numbers only (no answers disclosed)
    lines = grade_text.split("\n")
    in_breakdown = False
    missed_qs = []
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("QUESTION BREAKDOWN:"):
            in_breakdown = True
            continue
        if in_breakdown and stripped and stripped[0].isupper() and ":" in stripped and not stripped.startswith("Q"):
            break
        if in_breakdown and stripped.startswith("Q"):
            m = re.match(r"(Q\d+)\s*\([^)]*\):\s*(\d+)/(\d+)", stripped)
            if m and int(m.group(2)) < int(m.group(3)):
                missed_qs.append(m.group(1))

    if missed_qs:
        parts.append(f"Missed: {', '.join(missed_qs)}")

    # General comments (strip FR-related lines)
    comments = extract_comments(grade_text)
    if comments:
        safe_lines = []
        for line in comments.split("\n"):
            stripped = line.strip()
            if stripped.startswith(("FR", "Free Response", "Rubric")):
                continue
            safe_lines.append(line)
        safe = "\n".join(safe_lines).strip()
        if safe:
            parts.append(safe)

    return "\n".join(parts)


def extract_comments(grade_text: str) -> str:
    """Extract the COMMENTS section from grade output."""
    lines = grade_text.split("\n")
    in_comments = False
    comment_lines = []

    for line in lines:
        if line.strip().startswith("COMMENTS:"):
            in_comments = True
            continue
        if in_comments and line.strip().startswith(("SCAN/FORMAT", "DEDUCTIONS:")):
            break
        if in_comments:
            comment_lines.append(line)

    return "\n".join(comment_lines).strip() if comment_lines else ""


def read_grades_from_excel(excel_path: str) -> list[dict]:
    """
    Read an enriched grading Excel file and return grade records.

    Returns list of dicts:
        {name, score, total, questions_missed, comments}
    """
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True)
    ws = wb["Grades"]

    # Find columns by header text (resilient to reordering)
    header_map = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value:
            header_map[cell.value.strip()] = cell.column - 1  # 0-indexed

    col_name = header_map.get("Student", 0)
    col_score = header_map.get("Raw Score", 1)
    col_total = header_map.get("Total Points", 2)
    col_missed = header_map.get("Questions Missed")
    col_comments = header_map.get("Comments")
    col_canvas_comment = header_map.get("Canvas Comment")
    col_flags = header_map.get("Flags")

    records = []
    all_cols = [col_name, col_score, col_total]
    for c in (col_missed, col_comments, col_canvas_comment, col_flags):
        if c is not None:
            all_cols.append(c)
    max_col_needed = max(all_cols)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= col_name or not row[col_name]:
            continue

        if len(row) <= max_col_needed:
            logger.warning("Skipping short row (len=%d): %s", len(row), row)
            continue

        score_val = row[col_score]
        if score_val in ("ERROR", "N/A", None):
            name = row[col_name] if len(row) > col_name else "?"
            logger.warning("Skipping '%s' — score is %s", name, score_val)
            continue

        try:
            score = float(score_val)
        except (ValueError, TypeError):
            name = row[col_name] if len(row) > col_name else "?"
            logger.warning("Skipping '%s' — invalid score: %s", name, score_val)
            continue

        total = float(row[col_total]) if row[col_total] else None
        missed = str(row[col_missed]).strip() if col_missed is not None and row[col_missed] else ""
        comments = str(row[col_comments]).strip() if col_comments is not None and row[col_comments] else ""
        canvas_comment = str(row[col_canvas_comment]).strip() if col_canvas_comment is not None and len(row) > col_canvas_comment and row[col_canvas_comment] else ""
        flags = str(row[col_flags]).strip() if col_flags is not None and len(row) > col_flags and row[col_flags] else ""

        records.append({
            "name": str(row[col_name]).strip(),
            "score": score,
            "total": total,
            "questions_missed": missed,
            "comments": comments,
            "canvas_comment": canvas_comment,
            "flags": flags,
        })

    wb.close()
    return records


def _normalize_name(name: str) -> str:
    """Normalize a name for matching: lowercase, strip, flip 'Last, First' -> 'First Last'."""
    name = name.lower().strip()
    # Remove common PDF filename artifacts
    name = re.sub(r"_(?:submission|lab\d*|exam\d*|test\d*).*$", "", name, flags=re.IGNORECASE)
    name = name.replace("_", " ").replace(".", " ")
    # Flip "Last, First" -> "First Last"
    if "," in name:
        parts = [p.strip() for p in name.split(",", 1)]
        name = f"{parts[1]} {parts[0]}"
    return " ".join(name.split())  # collapse whitespace


def match_students_to_canvas(
    grade_records: list[dict],
    canvas_students: list[dict],
    threshold: float = 0.85,
) -> list[dict]:
    """
    Fuzzy-match student names from Excel to Canvas roster.

    Returns enriched list with canvas_student_id and canvas_match_name added.
    Unmatched students get canvas_student_id=None.
    """
    # Build normalized lookup from Canvas roster
    canvas_lookup = []
    for s in canvas_students:
        sid = s["id"]
        forms = set()
        forms.add(_normalize_name(s.get("name", "")))
        forms.add(_normalize_name(s.get("sortable_name", "")))
        if s.get("short_name"):
            forms.add(_normalize_name(s["short_name"]))
        forms.discard("")
        canvas_lookup.append({"id": sid, "name": s.get("name", ""), "forms": forms})

    matched = []
    for record in grade_records:
        # Build candidate names: primary + any alternate readings from handwriting OCR
        candidate_norms = [_normalize_name(record["name"])]
        for alt in record.get("name_alternates", []):
            n = _normalize_name(alt)
            if n and n not in candidate_norms:
                candidate_norms.append(n)

        best_id = None
        best_name = None
        best_score = 0.0
        second_best_score = 0.0

        for norm in candidate_norms:
            for entry in canvas_lookup:
                # Exact match on any name form
                if norm in entry["forms"]:
                    best_id = entry["id"]
                    best_name = entry["name"]
                    best_score = 1.0
                    break

                # Fuzzy match against all forms
                for form in entry["forms"]:
                    ratio = SequenceMatcher(None, norm, form).ratio()
                    if ratio > best_score:
                        second_best_score = best_score
                        best_score = ratio
                        best_id = entry["id"]
                        best_name = entry["name"]
                    elif ratio > second_best_score:
                        second_best_score = ratio
            if best_score == 1.0:
                break  # exact match found, no need to try alternates

        result = {**record, "match_confidence": round(best_score, 2)}
        if best_score >= threshold:
            result["canvas_student_id"] = best_id
            result["canvas_match_name"] = best_name
            # Flag near-miss ambiguity: two Canvas students with similar match scores
            if second_best_score >= threshold - 0.05:
                result["match_ambiguous"] = True
                logger.warning(
                    "AMBIGUOUS MATCH: '%s' matched '%s' (%.2f) but another student scored %.2f. Verify this match manually.",
                    record["name"], best_name, best_score, second_best_score,
                )
            else:
                result["match_ambiguous"] = False
        else:
            result["canvas_student_id"] = None
            result["canvas_match_name"] = None
            result["match_ambiguous"] = False

        matched.append(result)

    # Duplicate Canvas ID detection — abort if same student matched twice
    id_counts: dict[int, list[str]] = {}
    for m in matched:
        cid = m.get("canvas_student_id")
        if cid is not None:
            id_counts.setdefault(cid, []).append(m["name"])
    duplicates = {cid: names for cid, names in id_counts.items() if len(names) > 1}
    if duplicates:
        dup_details = "; ".join(
            f"Canvas ID {cid}: {', '.join(names)}" for cid, names in duplicates.items()
        )
        raise ValueError(
            f"DUPLICATE CANVAS ID MATCHES DETECTED — cannot post grades safely. "
            f"Multiple Excel students matched to the same Canvas student: {dup_details}. "
            f"Fix the Excel names or use a higher match threshold."
        )

    return matched


def format_canvas_comment(questions_missed: str) -> str:
    """
    Format a Canvas comment with only MC question numbers (no answers disclosed).
    Exams are reused — never include answer content or reasons.

    Args:
        questions_missed: Full detail string like "Q1 (2/3): binary fission missing; Q5 (0/1): wrote cell growth"

    Returns:
        Safe comment like "Missed: Q1, Q5" or "" if nothing missed.
    """
    if not questions_missed:
        return ""

    # Extract just question numbers: Q1, Q2, etc.
    q_numbers = re.findall(r"(Q\d+)", questions_missed)
    if not q_numbers:
        return ""

    # Deduplicate while preserving order
    seen = set()
    unique = []
    for q in q_numbers:
        if q not in seen:
            seen.add(q)
            unique.append(q)

    return f"Missed: {', '.join(unique)}"
