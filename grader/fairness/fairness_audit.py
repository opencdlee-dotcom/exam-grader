"""
Fairness Auditing Module

Detects potential disparities in grading across groups (sections, demographics).

Key Metric: Disparate Impact Ratio (EEOC Four-Fifths / 80% Rule)
- DI = min(mean_a, mean_b) / max(mean_a, mean_b)
- DI = 1.0: perfect parity
- DI ≥ 0.80: PASS — within the legal/ethical threshold (EEOC four-fifths rule)
- DI 0.65–0.79: WARNING — meaningful disparity, investigate before posting
- DI < 0.65: ALERT — substantial disparity, do not post without review

Threshold rationale (updated to EEOC 80% rule):
- The previous PASS threshold of 0.95 was too strict, flagging acceptable variation
  as warnings and desensitizing instructors to real alerts
- 0.80 is the EEOC four-fifths rule — the legally recognized minimum acceptable
  selection ratio, widely adopted in educational fairness auditing
- Source: EEOC Uniform Guidelines on Employee Selection Procedures (adapted for education)

Current v1: Section-level analysis only (T5 vs W2 vs W5)
Future v2: Demographic analysis (requires enrollment data)
"""

from typing import Dict, List, Tuple
import statistics
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def disparate_impact_ratio(
    scores_a: List[float], scores_b: List[float], min_n: int = 3
) -> Tuple[float, str]:
    """
    Calculate Disparate Impact (DI) ratio between two groups.

    DI = min(mean_a, mean_b) / max(mean_a, mean_b)

    Formula ensures DI ≤ 1.0, where 1.0 = perfect parity.

    Args:
        scores_a: Scores from group A
        scores_b: Scores from group B
        min_n: Minimum group size to calculate (default 3)

    Returns:
        Tuple of (di_ratio, interpretation_string)
    """
    if not scores_a or not scores_b or len(scores_a) < min_n or len(scores_b) < min_n:
        return 1.0, "INSUFFICIENT_DATA"

    mean_a = statistics.mean(scores_a)
    mean_b = statistics.mean(scores_b)

    if mean_a == 0 and mean_b == 0:
        return 1.0, "BOTH_ZERO"

    di = min(mean_a, mean_b) / max(mean_a, mean_b) if max(mean_a, mean_b) > 0 else 0.0

    # Interpret — EEOC four-fifths (80%) rule as primary threshold
    if di >= 0.80:
        status = "PASS"    # Within legal/ethical threshold
    elif di >= 0.65:
        status = "WARNING"  # Meaningful disparity — investigate
    else:
        status = "ALERT"   # Substantial disparity — do not post without review

    return round(di, 3), status


def analyze_section_grades(
    results: List[Dict], section_key: str = "section"
) -> Dict[str, Dict]:
    """
    Analyze grade distribution by section.

    Args:
        results: List of student result dicts with 'score' and section info
        section_key: Key in result dict for section identifier (default: "section")

    Returns:
        Dict mapping section_name -> {count, mean, stdev, min, max, di}
    """
    sections = {}

    # Group by section
    for result in results:
        section = result.get(section_key, "Unknown")
        if section not in sections:
            sections[section] = []
        sections[section].append(result.get("score", 0))

    # Calculate stats for each section
    analysis = {}
    for section_name, scores in sections.items():
        if not scores:
            continue

        analysis[section_name] = {
            "count": len(scores),
            "mean": round(statistics.mean(scores), 1),
            "stdev": round(statistics.stdev(scores), 1) if len(scores) > 1 else 0.0,
            "min": min(scores),
            "max": max(scores),
            "median": statistics.median(scores),
        }

    # Calculate DI between sections (pairwise)
    section_names = list(analysis.keys())
    for i, sec_a in enumerate(section_names):
        for sec_b in section_names[i + 1 :]:
            di, status = disparate_impact_ratio(sections[sec_a], sections[sec_b])
            analysis[sec_a][f"di_vs_{sec_b}"] = di
            analysis[sec_b][f"di_vs_{sec_a}"] = di

    return analysis


def create_fairness_audit_sheet(
    results: List[Dict], excel_path: str, section_key: str = "section"
) -> None:
    """
    Create Excel sheet with fairness audit data.

    Args:
        results: Student results
        excel_path: Path to Excel file
        section_key: Key for section identifier
    """
    wb = openpyxl.load_workbook(excel_path)

    # Remove old sheet if exists
    if "Fairness Audit" in wb.sheetnames:
        del wb["Fairness Audit"]

    # Create new sheet
    ws = wb.create_sheet("Fairness Audit", index=1)

    # Headers
    headers = ["Section", "Count", "Mean Score", "Std Dev", "Min", "Max", "Median", "DI Status"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Get analysis
    analysis = analyze_section_grades(results, section_key)

    # Write data
    for row_num, (section_name, stats) in enumerate(sorted(analysis.items()), start=2):
        # Determine DI status
        # EEOC four-fifths rule thresholds (consistent with disparate_impact_ratio)
        di_status = "PASS"
        for key in stats:
            if key.startswith("di_vs_"):
                if stats[key] < 0.80:
                    di_status = "WARNING"
                if stats[key] < 0.65:
                    di_status = "ALERT"

        row_data = [
            section_name,
            stats["count"],
            stats["mean"],
            stats["stdev"],
            stats["min"],
            stats["max"],
            stats["median"],
            di_status,
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="center")

            # Color code DI status
            if col_num == len(row_data):  # Last column (DI Status)
                if di_status == "PASS":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100", bold=True)
                elif di_status == "WARNING":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(color="9C6500", bold=True)
                else:  # ALERT
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006", bold=True)

    # Auto-width columns
    for col_num in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=row_num, column=col_num).value))
            for row_num in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max_length + 2

    wb.save(excel_path)


def generate_fairness_report(results: List[Dict]) -> str:
    """
    Generate human-readable fairness audit report.

    Args:
        results: Student results

    Returns:
        Formatted report string
    """
    analysis = analyze_section_grades(results)

    report_lines = ["=== FAIRNESS AUDIT REPORT ===\n"]

    if len(analysis) < 2:
        report_lines.append("✓ Single section only; no cross-section comparison")
        return "\n".join(report_lines)

    # Summary by section
    report_lines.append("SECTION SUMMARY")
    report_lines.append("─" * 60)
    report_lines.append("Section      | N  | Mean  | Std | Min | Max | Status")
    report_lines.append("─" * 60)

    for section_name in sorted(analysis.keys()):
        stats = analysis[section_name]
        status = "✓"
        for key in stats:
            if key.startswith("di_vs_"):
                if stats[key] < 0.80:  # EEOC four-fifths threshold
                    status = "⚠"
        report_lines.append(
            f"{section_name:12} | {stats['count']:2} | {stats['mean']:5.1f} | {stats['stdev']:3.1f} | "
            f"{stats['min']:3.0f} | {stats['max']:3.0f} | {status}"
        )

    report_lines.append("\nDISPARITE IMPACT ANALYSIS")
    report_lines.append("─" * 60)

    section_names = sorted(analysis.keys())
    any_disparity = False
    for i, sec_a in enumerate(section_names):
        for sec_b in section_names[i + 1 :]:
            di, status = disparate_impact_ratio(
                [r["score"] for r in results if r.get("section") == sec_a],
                [r["score"] for r in results if r.get("section") == sec_b],
            )
            if status != "PASS":
                any_disparity = True
            report_lines.append(f"{sec_a} vs {sec_b:12}: DI = {di:.3f}  [{status}]")

    if not any_disparity:
        report_lines.append("\n✓ No significant disparities detected")
    else:
        report_lines.append("\n⚠ Disparities detected. Investigate potential causes:")
        report_lines.append("  • Question difficulty variation by section")
        report_lines.append("  • Student prior knowledge differences")
        report_lines.append("  • Time-of-day testing effects")
        report_lines.append("  • Grading inconsistency (recalibrate rubric)")

    return "\n".join(report_lines)
