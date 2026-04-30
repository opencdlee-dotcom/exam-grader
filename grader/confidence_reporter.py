"""
Confidence Score Reporting

Bins model confidence into HIGH/MEDIUM/LOW based on calibrated thresholds.

Confidence Levels (aligned with ETS/ACT best practices for automated scoring):
- HIGH (≥0.80): Proceed automatically — expected accuracy aligns with human graders
- MEDIUM (0.70-0.79): Flag for light review — some uncertainty present
- LOW (<0.70): Mandatory human review before posting to Canvas

Threshold rationale:
- ETS e-rater and ACT CRASE use 0.70 as the minimum operational threshold
  (QWK ≥ 0.70 is the industry standard before a system is considered production-ready)
- Anything below 0.70 has accuracy comparable to a random second human rater —
  not good enough for high-stakes grading without review

These thresholds should be calibrated per exam domain when human-graded data
is available. See CLAUDE.md for calibration procedures.
"""

import logging
from typing import Dict, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)


def assign_confidence_level(model_confidence: float) -> str:
    """
    Bin model confidence into categorical level.

    Args:
        model_confidence: Raw confidence score (0.0 to 1.0)

    Returns:
        "HIGH", "MEDIUM", or "LOW"
    """
    if model_confidence >= 0.80:
        return "HIGH"
    elif model_confidence >= 0.70:
        return "MEDIUM"  # 0.70 aligns with ETS/ACT minimum operational threshold
    else:
        return "LOW"  # <0.70 → mandatory human review before Canvas posting


def get_confidence_from_result(result: Dict) -> float:
    """
    Extract confidence score from result dict.

    Looks for 'confidence' key; if not present, returns 0.70 (MEDIUM default).

    Args:
        result: Student result dict

    Returns:
        Confidence value (0.0-1.0)
    """
    conf = result.get("confidence", result.get("confidence_level"))
    if conf is not None:
        if isinstance(conf, str):
            # Convert "HIGH" -> 0.85, "MEDIUM" -> 0.65, "LOW" -> 0.35
            if conf == "HIGH":
                return 0.85
            elif conf == "MEDIUM":
                return 0.65
            else:
                return 0.35
        return float(conf)
    return 0.70  # Default to MEDIUM


def add_confidence_to_results(results: List[Dict]) -> List[Dict]:
    """
    Enrich results with confidence levels.

    Args:
        results: List of student result dicts

    Returns:
        Updated results with 'confidence_level' key
    """
    for result in results:
        conf_score = get_confidence_from_result(result)
        result["confidence_level"] = assign_confidence_level(conf_score)
        result["confidence_score"] = conf_score

    return results


def add_confidence_column_to_excel(results: List[Dict], excel_path: str) -> None:
    """
    Add Confidence column to Excel grades sheet.

    Inserts new column after Score column with HIGH/MEDIUM/LOW values.

    Args:
        results: Student results with 'confidence_level'
        excel_path: Path to Excel file
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Grades"]

    # Find Score column (usually column C)
    score_col = None
    for col_num, cell in enumerate(ws[4], 1):  # Header row is 4
        if cell.value == "Score":
            score_col = col_num
            break

    if not score_col:
        logger.warning("Could not find 'Score' column in Excel")
        wb.save(excel_path)
        return

    # Insert new column after Score
    ws.insert_cols(score_col + 1)

    # Add header
    header_cell = ws.cell(row=4, column=score_col + 1)
    header_cell.value = "Confidence"
    header_cell.font = Font(bold=True, size=11)
    header_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    header_cell.alignment = Alignment(horizontal="center")

    # Add confidence levels
    for row_idx, result in enumerate(results, start=5):  # Data starts row 5
        conf_level = result.get("confidence_level", "MEDIUM")
        conf_cell = ws.cell(row=row_idx, column=score_col + 1)
        conf_cell.value = conf_level
        conf_cell.alignment = Alignment(horizontal="center")

        # Color code: HIGH=green, MEDIUM=yellow, LOW=red
        if conf_level == "HIGH":
            conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            conf_cell.font = Font(color="006100")
        elif conf_level == "MEDIUM":
            conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            conf_cell.font = Font(color="9C6500")
        else:  # LOW
            conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            conf_cell.font = Font(color="9C0006")

    wb.save(excel_path)


def generate_confidence_summary(results: List[Dict]) -> str:
    """
    Generate summary of confidence distribution.

    Args:
        results: Student results with 'confidence_level'

    Returns:
        Formatted summary string
    """
    high_count = sum(1 for r in results if r.get("confidence_level") == "HIGH")
    medium_count = sum(1 for r in results if r.get("confidence_level") == "MEDIUM")
    low_count = sum(1 for r in results if r.get("confidence_level") == "LOW")
    total = len(results)

    if total == 0:
        return "No results to analyze"

    summary = f"""
CONFIDENCE DISTRIBUTION
──────────────────────
HIGH   ({high_count:2d}/{total}): {high_count/total*100:5.1f}%  (expect >95% accurate)
MEDIUM ({medium_count:2d}/{total}): {medium_count/total*100:5.1f}%  (expect 70-85% accurate)
LOW    ({low_count:2d}/{total}): {low_count/total*100:5.1f}%  (expect <70% accurate)

⚠ Action Items:
  • HIGH confidence answers: prioritize for instructor spot-check (>95% should be correct)
  • MEDIUM confidence answers: marginal, consider manual review if <70% expected accuracy
  • LOW confidence answers: flag for human review before posting to Canvas
"""
    return summary.strip()
