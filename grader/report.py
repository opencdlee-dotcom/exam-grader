"""
Generate grade reports from scoring results.
Outputs to console and/or file.

Excel output includes:
- Grades sheet: Student, Raw Score, Total Points, %, Questions Missed, Comments, Flags, FR Details (Private), Canvas Comment
- Flagged Students sheet: Filtered view of all flagged students with details
- Answer Key sheet: Q#, Type, Points, Answer
- Statistics sheet: Class stats with quartiles
"""

import logging
import os
import re
import statistics
from datetime import datetime

logger = logging.getLogger(__name__)


def extract_score(grade_text: str) -> tuple[float | None, float | None]:
    """Extract numeric score from grade text. Returns (score, total) or (None, None).

    Tries multiple patterns in priority order to handle variations in Claude's output format
    (with/without spaces, different label words).
    """
    patterns = [
        r"GRADE\s*[:\s]\s*([\d.]+)\s*/\s*([\d.]+)",
        r"TOTAL\s*[:\s]\s*([\d.]+)\s*/\s*([\d.]+)",
        r"SCORE\s*[:\s]\s*([\d.]+)\s*/\s*([\d.]+)",
    ]
    for pat in patterns:
        match = re.search(pat, grade_text, re.IGNORECASE)
        if match:
            return float(match.group(1)), float(match.group(2))
    return None, None


def extract_score_structured(result: dict) -> tuple[float | None, float | None]:
    """Extract score from structured result, falling back to regex.
    Sets _score_source on the result dict for downstream tracking."""
    sr = result.get("structured_result")
    if sr and hasattr(sr, 'total_score') and sr.total_score is not None:
        result["_score_source"] = "structured"
        return sr.total_score, sr.total_possible
    result["_score_source"] = "regex_fallback"
    return extract_score(result.get("grade_text", ""))


def _extract_missed_questions(grade_text: str) -> str:
    """
    Extract missed questions from grade text.

    Handles two formats:
    - Exam mode: QUESTION BREAKDOWN section with Q1 (3 pts): 2/3 — reason
    - Notebook mode: DEDUCTIONS section with -N reason lines

    Returns a string describing what was missed, or "" if nothing found.
    """
    lines = grade_text.split("\n")

    # Try exam mode: QUESTION BREAKDOWN section
    in_breakdown = False
    missed = []
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("QUESTION BREAKDOWN:"):
            in_breakdown = True
            continue
        if in_breakdown and stripped and stripped[0].isupper() and ":" in stripped and not stripped.startswith("Q"):
            break  # Hit next section header
        if in_breakdown and stripped.startswith("Q"):
            # Parse Q1 (3 pts): 2/3 — reason
            m = re.match(r"(Q\d+)\s*\([^)]*\):\s*(\d+)/(\d+)\s*[—–-]\s*(.*)", stripped)
            if m:
                q_num, earned, total = m.group(1), int(m.group(2)), int(m.group(3))
                reason = m.group(4).strip()
                if earned < total:
                    missed.append(f"{q_num} ({earned}/{total}): {reason}")

    if missed:
        return "; ".join(missed)

    # Try notebook mode: DEDUCTIONS section
    in_deductions = False
    deductions = []
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("DEDUCTIONS:"):
            in_deductions = True
            continue
        if in_deductions and stripped and stripped[0].isupper() and ":" in stripped and not stripped.startswith("-"):
            break  # Hit next section header
        if in_deductions and stripped.startswith("-"):
            deductions.append(stripped)

    if deductions:
        return "; ".join(deductions)

    return ""


def _extract_comments_local(grade_text: str) -> str:
    """Extract the COMMENTS section from grade output."""
    lines = grade_text.split("\n")
    in_comments = False
    comment_lines = []

    for line in lines:
        if line.strip().startswith("COMMENTS:"):
            in_comments = True
            continue
        if in_comments and line.strip().startswith(("SCAN/FORMAT", "DEDUCTIONS:")):
            break
        if in_comments:
            comment_lines.append(line)

    return "\n".join(comment_lines).strip() if comment_lines else ""


def _extract_flags(result: dict) -> str:
    """
    Build consolidated flags string from structured result.
    Returns flags like: "⚠ BELOW 50%, 🔍 HUMAN REVIEW (Q3, Q7)"
    """
    flags = []
    sr = result.get("structured_result")

    # Below 50%
    score, total = extract_score_structured(result)
    if score is not None and total and total > 0:
        if (score / total) < 0.5:
            flags.append("⚠ BELOW 50%")

    if sr:
        # Human review needed
        if hasattr(sr, 'needs_human_review') and sr.needs_human_review:
            low_items = []
            if hasattr(sr, 'low_confidence_items'):
                for item in sr.low_confidence_items:
                    # Extract Q number if present
                    m = re.match(r"(Q\d+|FR\d+)", item)
                    if m:
                        low_items.append(m.group(1))
            detail = f" ({', '.join(low_items)})" if low_items else ""
            flags.append(f"🔍 HUMAN REVIEW{detail}")

        # Ambiguous questions (LOW confidence — includes illegible)
        ambiguous = []
        for q in getattr(sr, 'question_results', []):
            if q.confidence.value == "LOW":
                ambiguous.append(f"Q{q.question_number}")
        if ambiguous:
            flags.append(f"❓ AMBIGUOUS ({', '.join(ambiguous)})")

        # Shift detected
        if getattr(sr, 'shift_detected', False):
            flags.append("📝 SHIFT DETECTED")

        # Faint content
        faint_qs = [f"Q{q.question_number}" for q in getattr(sr, 'question_results', []) if q.faint_content]
        if faint_qs:
            flags.append(f"👁 FAINT CONTENT ({', '.join(faint_qs)})")

        # Recheck result
        if result.get("recheck_result"):
            rc = result["recheck_result"]
            flags.append(f"🔄 RECHECKED ({rc.get('original_score', '?')}→{rc.get('final_score', '?')})")

    # Score extraction degraded to regex fallback
    if result.get("_score_source") == "regex_fallback":
        flags.append("⚙ REGEX FALLBACK (structured parse failed)")

    return ", ".join(flags)


def _extract_fr_details(result: dict) -> str:
    """
    Extract full free response breakdown with reasoning (instructor-only).
    Never post this to Canvas.
    """
    sr = result.get("structured_result")
    if not sr:
        # Fall back to text parsing
        grade_text = result.get("grade_text", "")
        lines = grade_text.split("\n")
        in_fr = False
        fr_lines = []
        for line in lines:
            if line.strip().startswith("FREE RESPONSE BREAKDOWN:"):
                in_fr = True
                continue
            if in_fr and line.strip().startswith(("COMMENTS:", "CONFIDENCE", "SCAN/FORMAT")):
                break
            if in_fr:
                fr_lines.append(line)
        return "\n".join(fr_lines).strip()

    fr_results = getattr(sr, 'free_response_results', [])
    if not fr_results:
        return ""

    lines = []
    for fr in fr_results:
        conf = fr.confidence.value
        lines.append(
            f"FR{fr.question_number} ({fr.points_possible:.0f} pts): "
            f"{fr.points_earned:.0f}/{fr.points_possible:.0f} [{conf}] — {fr.reasoning}"
        )
    return "\n".join(lines)


def _build_canvas_comment(result: dict) -> str:
    """
    Build a safe Canvas comment: missed question numbers + general feedback.
    Never includes FR details, specific wrong answers, or rubric reasoning.
    """
    parts = []

    # Missed questions (numbers only, no answers)
    grade_text = result.get("grade_text", "")
    missed = _extract_missed_questions(grade_text)
    if missed:
        from canvas.grades import format_canvas_comment
        safe_missed = format_canvas_comment(missed)
        if safe_missed:
            parts.append(safe_missed)

    # General comments (feedback, not answer content)
    comments = _extract_comments_local(grade_text)
    if comments:
        # Strip any lines that look like they contain answer content
        safe_lines = []
        for line in comments.split("\n"):
            stripped = line.strip()
            # Skip lines that reference specific answers or FR reasoning
            if stripped.startswith(("FR", "Free Response", "Rubric")):
                continue
            safe_lines.append(line)
        safe_comment = "\n".join(safe_lines).strip()
        if safe_comment:
            parts.append(safe_comment)

    return "\n".join(parts)


def print_result(result: dict):
    """Log a single grading result."""
    logger.info("=" * 60)
    logger.info("Student: %s", result["student_file"])
    logger.info("Assignment: %s", result.get("answer_key_title", "N/A"))
    logger.info("=" * 60)
    logger.info("%s", result["grade_text"])


def save_results(results: list[dict], output_path: str = None) -> str:
    """
    Save all grading results to a markdown file.

    Returns:
        Path to the saved file
    """
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        os.makedirs(os.path.join("output", "grades"), exist_ok=True)
        output_path = os.path.join("output", "grades", f"grades_{timestamp}.md")

    lines = [
        f"# Grading Results",
        f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
    ]

    # Summary table
    lines.append("## Summary")
    lines.append("")
    lines.append("| Student | Score | Status |")
    lines.append("|---------|-------|--------|")

    for result in results:
        name = result["student_file"]
        if result.get("error"):
            lines.append(f"| {name} | ERROR | Failed |")
        else:
            score, total = extract_score(result["grade_text"])
            if score is not None:
                lines.append(f"| {name} | {score}/{total} | Graded |")
            else:
                lines.append(f"| {name} | ? | Check output |")

    lines.append("")

    # Detailed results
    lines.append("## Detailed Results")
    for result in results:
        lines.append(f"\n### {result['student_file']}")
        lines.append("")
        lines.append(result["grade_text"])
        lines.append("")

    with open(output_path, "w") as f:
        f.write("\n".join(lines))

    return output_path


def print_summary_table(posted_results: list[dict], total_points: float = None):
    """
    Print a summary table of all graded students (notebook/test grader style).

    Args:
        posted_results: List of dicts with student_name, score, posted, error (optional)
        total_points: Total possible points for display
    """
    display_total = f"/{int(total_points)}" if total_points else ""

    logger.info("=" * 56)
    logger.info("%-25s | %8s | Status", "Student Name", "Score")
    logger.info("%s-+-%s-+-%s", "-" * 25, "-" * 8, "-" * 16)

    scores = []
    errors = 0
    for r in posted_results:
        name = r.get("student_name", "Unknown")[:25]
        score = r.get("score")
        if r.get("error"):
            logger.info("%-25s | %8s | ERROR: %s", name, "--", r["error"])
            errors += 1
        elif r.get("posted"):
            logger.info("%-25s | %s%s | Submitted", name, score, display_total)
            if score is not None:
                scores.append(score)
        else:
            logger.info("%-25s | %s%s | Skipped", name, score, display_total)

    logger.info("=" * 56)

    graded_count = len(scores)
    avg = sum(scores) / graded_count if graded_count > 0 else 0
    logger.info("TOTAL: %d graded | %d errors | AVG: %.1f%s", graded_count, errors, avg, display_total)


def save_results_excel(results: list[dict], output_path: str = None, answer_key: dict = None) -> str:
    """
    Save grading results to an Excel file with grades, flags, and statistics.

    Sheets:
        - Grades: Full breakdown per student (includes private FR details)
        - Flagged Students: Filtered view of flagged students
        - Answer Key: Reference (if answer_key provided)
        - Statistics: Class stats with quartiles

    Args:
        results: List of grade result dicts (from grade_batch or grade_exam_batch)
        output_path: Optional output path. Defaults to output/grades_TIMESTAMP.xlsx
        answer_key: Optional extracted answer key dict for the Answer Key sheet

    Returns:
        Path to the saved .xlsx file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        os.makedirs(os.path.join("output", "grades"), exist_ok=True)
        output_path = os.path.join("output", "grades", f"grades_{timestamp}.xlsx")

    wb = Workbook()

    # --- Shared styles ---
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_text = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap_align = Alignment(wrap_text=True, vertical="top")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    # === Grades sheet ===
    ws = wb.active
    ws.title = "Grades"

    headers = [
        "Student", "Raw Score", "Total Points", "Percentage",
        "Questions Missed", "Comments", "Flags",
        "FR Details (Private)", "Canvas Comment",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_text
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    scores = []
    total_points = None
    flagged_rows = []  # Collect for Flagged Students sheet
    row = 2

    for result in results:
        name = result.get("student_file", "Unknown")

        if result.get("error"):
            ws.cell(row=row, column=1, value=name).border = thin_border
            ws.cell(row=row, column=2, value="ERROR").border = thin_border
            for c in range(3, len(headers) + 1):
                ws.cell(row=row, column=c, value="").border = thin_border
            row += 1
            continue

        grade_text = result["grade_text"]
        score, total = extract_score_structured(result)

        # Col A: Student
        ws.cell(row=row, column=1, value=name).border = thin_border

        # Col B-D: Score, Total, Percentage
        pct = 0.0
        if score is not None:
            ws.cell(row=row, column=2, value=score).border = thin_border
            ws.cell(row=row, column=3, value=total).border = thin_border
            pct = round((score / total) * 100, 1) if total else 0
            cell = ws.cell(row=row, column=4, value=pct)
            cell.border = thin_border
            cell.number_format = '0.0"%"'
            scores.append(score)
            if total_points is None:
                total_points = total
        else:
            ws.cell(row=row, column=2, value="N/A").border = thin_border
            ws.cell(row=row, column=3, value="").border = thin_border
            ws.cell(row=row, column=4, value="").border = thin_border

        # Col E: Questions Missed
        missed = _extract_missed_questions(grade_text)
        cell_missed = ws.cell(row=row, column=5, value=missed)
        cell_missed.border = thin_border
        cell_missed.alignment = wrap_align

        # Col F: Comments
        comments = _extract_comments_local(grade_text)
        cell_comments = ws.cell(row=row, column=6, value=comments)
        cell_comments.border = thin_border
        cell_comments.alignment = wrap_align

        # Col G: Flags
        flags = _extract_flags(result)
        cell_flags = ws.cell(row=row, column=7, value=flags)
        cell_flags.border = thin_border
        cell_flags.alignment = wrap_align

        # Col H: FR Details (Private — instructor only, never posted)
        fr_details = _extract_fr_details(result)
        cell_fr = ws.cell(row=row, column=8, value=fr_details)
        cell_fr.border = thin_border
        cell_fr.alignment = wrap_align

        # Col I: Canvas Comment (safe for posting — no answers or FR reasoning)
        canvas_comment = _build_canvas_comment(result)
        cell_canvas = ws.cell(row=row, column=9, value=canvas_comment)
        cell_canvas.border = thin_border
        cell_canvas.alignment = wrap_align

        # Conditional formatting: red for below 50%, yellow for flagged
        if pct < 50 and score is not None:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = red_fill
        elif flags:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = yellow_fill

        # Track flagged students for the Flagged sheet
        if flags:
            flagged_rows.append({
                "name": name, "score": score, "total": total,
                "pct": pct, "flags": flags,
                "details": "; ".join(
                    getattr(result.get("structured_result"), "low_confidence_items", [])[:5]
                ) if result.get("structured_result") else "",
            })

        row += 1

    # Column widths
    col_widths = {"A": 30, "B": 14, "C": 14, "D": 14, "E": 50, "F": 40, "G": 35, "H": 60, "I": 45}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # === Flagged Students sheet ===
    ws_flagged = wb.create_sheet("Flagged Students")
    flag_headers = ["Student", "Score", "Percentage", "Flags", "Review Details"]
    for col, header in enumerate(flag_headers, 1):
        cell = ws_flagged.cell(row=1, column=col, value=header)
        cell.font = header_text
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for i, fr in enumerate(flagged_rows, 2):
        ws_flagged.cell(row=i, column=1, value=fr["name"]).border = thin_border
        score_str = f"{fr['score']}/{fr['total']}" if fr["score"] is not None else "N/A"
        ws_flagged.cell(row=i, column=2, value=score_str).border = thin_border
        cell_pct = ws_flagged.cell(row=i, column=3, value=fr["pct"])
        cell_pct.border = thin_border
        cell_pct.number_format = '0.0"%"'
        ws_flagged.cell(row=i, column=4, value=fr["flags"]).border = thin_border
        ws_flagged.cell(row=i, column=4).alignment = wrap_align
        ws_flagged.cell(row=i, column=5, value=fr["details"]).border = thin_border
        ws_flagged.cell(row=i, column=5).alignment = wrap_align

        # Color code
        if fr["pct"] < 50:
            for c in range(1, 6):
                ws_flagged.cell(row=i, column=c).fill = red_fill
        else:
            for c in range(1, 6):
                ws_flagged.cell(row=i, column=c).fill = yellow_fill

    ws_flagged.column_dimensions["A"].width = 30
    ws_flagged.column_dimensions["B"].width = 14
    ws_flagged.column_dimensions["C"].width = 14
    ws_flagged.column_dimensions["D"].width = 40
    ws_flagged.column_dimensions["E"].width = 60

    if not flagged_rows:
        ws_flagged.cell(row=2, column=1, value="No flagged students").font = Font(italic=True)

    # === Answer Key sheet ===
    if answer_key:
        ws_key = wb.create_sheet("Answer Key")

        if "questions" in answer_key:
            # Exam mode: questions list with answers and alternatives
            key_headers = ["Q#", "Type", "Points", "Answer", "Alternatives"]
            for col, header in enumerate(key_headers, 1):
                cell = ws_key.cell(row=1, column=col, value=header)
                cell.font = header_text
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            key_row = 2
            for q in answer_key["questions"]:
                ws_key.cell(row=key_row, column=1, value=f"Q{q.get('number', key_row - 1)}").border = thin_border
                ws_key.cell(row=key_row, column=2, value=q.get("type", "objective")).border = thin_border
                ws_key.cell(row=key_row, column=3, value=q.get("points", 1)).border = thin_border
                answers = q.get("answers", [])
                ws_key.cell(row=key_row, column=4, value=", ".join(answers) if isinstance(answers, list) else str(answers)).border = thin_border
                alts = q.get("alternatives", {})
                alt_str = "; ".join(f"{k}: {', '.join(v) if isinstance(v, list) else v}" for k, v in alts.items()) if alts else ""
                ws_key.cell(row=key_row, column=5, value=alt_str).border = thin_border
                ws_key.cell(row=key_row, column=5).alignment = wrap_align
                key_row += 1

            ws_key.column_dimensions["A"].width = 8
            ws_key.column_dimensions["B"].width = 14
            ws_key.column_dimensions["C"].width = 10
            ws_key.column_dimensions["D"].width = 40
            ws_key.column_dimensions["E"].width = 50

        elif "sections" in answer_key:
            # Notebook mode: sections with criteria
            key_headers = ["Section", "Points", "Criteria", "Notes"]
            for col, header in enumerate(key_headers, 1):
                cell = ws_key.cell(row=1, column=col, value=header)
                cell.font = header_text
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

            key_row = 2
            for section in answer_key["sections"]:
                ws_key.cell(row=key_row, column=1, value=section.get("name", "")).border = thin_border
                ws_key.cell(row=key_row, column=2, value=section.get("points", 0)).border = thin_border
                criteria = section.get("criteria", [])
                criteria_str = "\n".join(
                    f"- {c.get('item', '')} ({c.get('points', '')} pts)"
                    for c in criteria
                ) if criteria else ""
                cell_crit = ws_key.cell(row=key_row, column=3, value=criteria_str)
                cell_crit.border = thin_border
                cell_crit.alignment = wrap_align
                ws_key.cell(row=key_row, column=4, value=section.get("notes", "")).border = thin_border
                ws_key.cell(row=key_row, column=4).alignment = wrap_align
                key_row += 1

            ws_key.column_dimensions["A"].width = 30
            ws_key.column_dimensions["B"].width = 10
            ws_key.column_dimensions["C"].width = 50
            ws_key.column_dimensions["D"].width = 40

        # Add title and total at the top if available
        title = answer_key.get("title", "")
        total = answer_key.get("total_points", "")
        if title or total:
            ws_key.insert_rows(1)
            ws_key.cell(row=1, column=1, value=f"{title}  —  {total} pts" if total else title).font = Font(bold=True, size=13)

    # === Statistics sheet ===
    ws2 = wb.create_sheet("Statistics")

    stat_header_font = Font(bold=True, size=14)
    stat_label_font = Font(bold=True, size=11)
    stat_value_font = Font(size=11)

    ws2.cell(row=1, column=1, value="Class Statistics").font = stat_header_font
    ws2.cell(row=2, column=1, value=f"Total Points Possible: {total_points or 'N/A'}").font = Font(size=11, italic=True)
    ws2.cell(row=3, column=1, value=f"Students Graded: {len(scores)}  |  Flagged: {len(flagged_rows)}").font = Font(size=11, italic=True)

    if scores:
        sorted_scores = sorted(scores)
        q1 = statistics.quantiles(sorted_scores, n=4)[0] if len(sorted_scores) >= 2 else sorted_scores[0]
        q3 = statistics.quantiles(sorted_scores, n=4)[2] if len(sorted_scores) >= 2 else sorted_scores[-1]

        below_50_count = sum(1 for s in scores if total_points and (s / total_points) < 0.5)

        stat_rows = [
            ("Highest", max(scores)),
            ("Lowest", min(scores)),
            ("Average", round(statistics.mean(scores), 2)),
            ("Median", round(statistics.median(scores), 2)),
            ("Upper Quartile (Q3)", round(q3, 2)),
            ("Lower Quartile (Q1)", round(q1, 2)),
        ]

        # Headers
        for col, header in enumerate(["Statistic", "Value", "Percentage"], 1):
            cell = ws2.cell(row=5, column=col, value=header)
            cell.font = header_text
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for i, (label, value) in enumerate(stat_rows, 6):
            ws2.cell(row=i, column=1, value=label).font = stat_label_font
            ws2.cell(row=i, column=1).border = thin_border
            ws2.cell(row=i, column=2, value=value).font = stat_value_font
            ws2.cell(row=i, column=2).border = thin_border
            ws2.cell(row=i, column=2).alignment = Alignment(horizontal="center")
            if total_points:
                pct = round((value / total_points) * 100, 1)
                cell = ws2.cell(row=i, column=3, value=pct)
                cell.font = stat_value_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = '0.0"%"'

        # Below-50% summary
        summary_row = 6 + len(stat_rows) + 1
        ws2.cell(row=summary_row, column=1, value=f"Below 50%: {below_50_count} student(s)").font = Font(bold=True, color="CC0000")
        ws2.cell(row=summary_row + 1, column=1, value=f"Flagged for Review: {len(flagged_rows)} student(s)").font = Font(bold=True, color="CC7700")
    else:
        ws2.cell(row=5, column=1, value="No valid scores to compute statistics.").font = Font(italic=True)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    wb.save(output_path)
    return output_path


class UpdateResult:
    """Result of an update_excel_student operation."""
    UPDATED = "updated"
    NOT_FOUND = "not_found"
    MATCH_TOO_WEAK = "match_too_weak"

    def __init__(self, status: str, match_score: float = 0.0, matched_name: str = ""):
        self.status = status
        self.match_score = match_score
        self.matched_name = matched_name

    def __bool__(self):
        return self.status == self.UPDATED


def update_excel_student(
    excel_path: str,
    student_name: str,
    updates: dict,
) -> 'UpdateResult':
    """
    Update a single student's row in an existing grading Excel file.
    Writes to an Audit Log sheet for grade versioning.

    Args:
        excel_path: Path to the Excel file
        student_name: Student name to match (fuzzy)
        updates: Dict of column_name -> new_value (e.g. {"Raw Score": 35, "Comments": "..."})

    Returns:
        UpdateResult with status, match_score, and matched_name
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from difflib import SequenceMatcher

    wb = load_workbook(excel_path)
    ws = wb["Grades"]

    # Build header map
    header_map = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value:
            header_map[cell.value.strip()] = cell.column

    # Find student row (fuzzy match)
    col_name = header_map.get("Student", 1)
    best_row = None
    best_score = 0.0
    best_matched_name = ""
    for row in ws.iter_rows(min_row=2):
        cell_val = row[col_name - 1].value
        if not cell_val:
            continue
        ratio = SequenceMatcher(None, student_name.lower(), str(cell_val).lower()).ratio()
        if ratio > best_score:
            best_score = ratio
            best_row = row[0].row
            best_matched_name = str(cell_val)

    if best_row is None:
        return UpdateResult(UpdateResult.NOT_FOUND)

    if best_score < 0.7:
        return UpdateResult(UpdateResult.MATCH_TOO_WEAK, match_score=best_score, matched_name=best_matched_name)

    # Create or find Audit Log sheet
    if "Audit Log" not in wb.sheetnames:
        ws_audit = wb.create_sheet("Audit Log")
        ws_audit.cell(row=1, column=1, value="Timestamp").font = Font(bold=True)
        ws_audit.cell(row=1, column=2, value="Student").font = Font(bold=True)
        ws_audit.cell(row=1, column=3, value="Field").font = Font(bold=True)
        ws_audit.cell(row=1, column=4, value="Old Value").font = Font(bold=True)
        ws_audit.cell(row=1, column=5, value="New Value").font = Font(bold=True)
        ws_audit.column_dimensions["A"].width = 22
        ws_audit.column_dimensions["B"].width = 25
        ws_audit.column_dimensions["C"].width = 18
        ws_audit.column_dimensions["D"].width = 25
        ws_audit.column_dimensions["E"].width = 25
    else:
        ws_audit = wb["Audit Log"]
    audit_row = ws_audit.max_row + 1
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Apply updates and log to audit trail
    for col_name_key, new_value in updates.items():
        col_idx = header_map.get(col_name_key)
        if col_idx:
            old_value = ws.cell(row=best_row, column=col_idx).value
            ws.cell(row=best_row, column=col_idx, value=new_value)
            # Write audit log entry
            ws_audit.cell(row=audit_row, column=1, value=timestamp)
            ws_audit.cell(row=audit_row, column=2, value=best_matched_name)
            ws_audit.cell(row=audit_row, column=3, value=col_name_key)
            ws_audit.cell(row=audit_row, column=4, value=str(old_value) if old_value is not None else "")
            ws_audit.cell(row=audit_row, column=5, value=str(new_value) if new_value is not None else "")
            audit_row += 1

    # Recalculate percentage if score was updated
    if "Raw Score" in updates:
        col_score = header_map.get("Raw Score")
        col_total = header_map.get("Total Points")
        col_pct = header_map.get("Percentage")
        if col_score and col_total and col_pct:
            score = ws.cell(row=best_row, column=col_score).value
            total = ws.cell(row=best_row, column=col_total).value
            if score is not None and total:
                pct = round((float(score) / float(total)) * 100, 1)
                ws.cell(row=best_row, column=col_pct, value=pct)

    # Regenerate statistics sheet
    _regenerate_stats_sheet(wb, header_map)

    wb.save(excel_path)
    return UpdateResult(UpdateResult.UPDATED, match_score=best_score, matched_name=best_matched_name)


def _regenerate_stats_sheet(wb, header_map: dict):
    """Recalculate the Statistics sheet from current Grades data."""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    ws_grades = wb["Grades"]
    col_score = header_map.get("Raw Score", 2)
    col_total = header_map.get("Total Points", 3)
    col_flags = header_map.get("Flags", 7)

    scores = []
    total_points = None
    flagged_count = 0

    for row in ws_grades.iter_rows(min_row=2, values_only=False):
        score_val = row[col_score - 1].value
        if score_val in ("ERROR", "N/A", None, ""):
            continue
        try:
            scores.append(float(score_val))
        except (ValueError, TypeError):
            continue
        if total_points is None:
            total_val = row[col_total - 1].value
            if total_val:
                try:
                    total_points = float(total_val)
                except (ValueError, TypeError):
                    pass
        flag_val = row[col_flags - 1].value if col_flags and len(row) >= col_flags else None
        if flag_val:
            flagged_count += 1

    # Clear and rebuild stats sheet
    if "Statistics" in wb.sheetnames:
        del wb["Statistics"]

    ws2 = wb.create_sheet("Statistics")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws2.cell(row=1, column=1, value="Class Statistics").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"Total Points Possible: {total_points or 'N/A'}").font = Font(size=11, italic=True)
    ws2.cell(row=3, column=1, value=f"Students Graded: {len(scores)}  |  Flagged: {flagged_count}").font = Font(size=11, italic=True)

    if scores:
        sorted_scores = sorted(scores)
        q1 = statistics.quantiles(sorted_scores, n=4)[0] if len(sorted_scores) >= 2 else sorted_scores[0]
        q3 = statistics.quantiles(sorted_scores, n=4)[2] if len(sorted_scores) >= 2 else sorted_scores[-1]

        stat_rows = [
            ("Highest", max(scores)),
            ("Lowest", min(scores)),
            ("Average", round(statistics.mean(scores), 2)),
            ("Median", round(statistics.median(scores), 2)),
            ("Upper Quartile (Q3)", round(q3, 2)),
            ("Lower Quartile (Q1)", round(q1, 2)),
        ]

        for col, header in enumerate(["Statistic", "Value", "Percentage"], 1):
            cell = ws2.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for i, (label, value) in enumerate(stat_rows, 6):
            ws2.cell(row=i, column=1, value=label).font = Font(bold=True, size=11)
            ws2.cell(row=i, column=1).border = thin_border
            ws2.cell(row=i, column=2, value=value).font = Font(size=11)
            ws2.cell(row=i, column=2).border = thin_border
            ws2.cell(row=i, column=2).alignment = Alignment(horizontal="center")
            if total_points:
                pct = round((value / total_points) * 100, 1)
                cell = ws2.cell(row=i, column=3, value=pct)
                cell.font = Font(size=11)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = '0.0"%"'

        below_50 = sum(1 for s in scores if total_points and (s / total_points) < 0.5)
        summary_row = 6 + len(stat_rows) + 1
        ws2.cell(row=summary_row, column=1, value=f"Below 50%: {below_50} student(s)").font = Font(bold=True, color="CC0000")
        ws2.cell(row=summary_row + 1, column=1, value=f"Flagged for Review: {flagged_count} student(s)").font = Font(bold=True, color="CC7700")

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
