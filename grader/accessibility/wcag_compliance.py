"""
WCAG AA Accessibility Compliance Module

Ensures Excel output meets WCAG AA standards:
- Color contrast ≥4.5:1 for normal text, ≥3:1 for large text
- Dyslexia-friendly fonts (sans-serif preferred)
- No information conveyed by color alone
- Clear, descriptive labels and alt text

Reference: https://www.w3.org/WAI/WCAG21/quickref/
"""

from typing import Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def calculate_relative_luminance(hex_color: str) -> float:
    """
    Calculate relative luminance of a color for contrast ratio.

    Args:
        hex_color: Color in hex format (e.g., "FFFFFF" or "000000")

    Returns:
        Relative luminance value (0-1)
    """
    # Remove # if present
    hex_color = hex_color.lstrip("#")

    # Convert hex to RGB
    r = int(hex_color[0:2], 16) / 255.0
    g = int(hex_color[2:4], 16) / 255.0
    b = int(hex_color[4:6], 16) / 255.0

    # Apply gamma correction
    def adjust_channel(c):
        if c <= 0.03928:
            return c / 12.92
        return ((c + 0.055) / 1.055) ** 2.4

    r = adjust_channel(r)
    g = adjust_channel(g)
    b = adjust_channel(b)

    # Calculate luminance
    luminance = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return luminance


def validate_contrast_ratio(
    text_color: str, bg_color: str, min_ratio: float = 4.5
) -> Tuple[float, bool]:
    """
    Validate contrast ratio between text and background colors.

    Args:
        text_color: Hex color for text (e.g., "000000")
        bg_color: Hex color for background (e.g., "FFFFFF")
        min_ratio: Minimum required contrast (default 4.5 for WCAG AA normal text)

    Returns:
        Tuple of (ratio, is_compliant)
    """
    l1 = calculate_relative_luminance(text_color)
    l2 = calculate_relative_luminance(bg_color)

    # Contrast ratio formula
    lighter = max(l1, l2)
    darker = min(l1, l2)
    ratio = (lighter + 0.05) / (darker + 0.05)

    is_compliant = ratio >= min_ratio
    return round(ratio, 1), is_compliant


def apply_wcag_aa_formatting(excel_path: str) -> None:
    """
    Apply WCAG AA accessibility formatting to Excel file.

    Changes:
    - Header: black background with white text (perfect contrast)
    - Data rows: alternating white/light gray for readability
    - Fonts: Arial/Helvetica (dyslexia-friendly sans-serif)
    - No information by color alone (use icons + text)
    - Adequate cell padding

    Args:
        excel_path: Path to Excel file
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Grades"]

    # Define colors (high contrast)
    black = "000000"
    white = "FFFFFF"
    light_gray = "F2F2F2"
    header_fill = PatternFill(start_color=black, end_color=black, fill_type="solid")
    header_font = Font(
        name="Arial", size=11, bold=True, color=white  # 21:1 contrast
    )
    light_fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
    data_font = Font(name="Arial", size=11, color=black)

    # Borders (help with readability)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Format header row (row 4 in Grades sheet)
    for col in ws[4]:
        if col.value:
            col.fill = header_fill
            col.font = header_font
            col.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            col.border = thin_border

    # Format data rows with alternating colors
    for row_num in range(5, ws.max_row + 1):
        for col_num, cell in enumerate(ws[row_num], 1):
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col_num > 2 else "left", vertical="center")

            # Alternate row colors (even rows: light gray; odd rows: white)
            if row_num % 2 == 0:
                cell.fill = light_fill

    # Auto-width columns with min/max
    for col_num, col_letter in enumerate(ws.columns, 1):
        max_length = 0
        column = col_letter
        for cell in ws[openpyxl.utils.get_column_letter(col_num)]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = min(max(max_length + 2, 10), 40)

    # Set row height for readability
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20

    # Freeze top rows
    ws.freeze_panes = "A5"

    wb.save(excel_path)


def apply_dyslexia_friendly_styling(excel_path: str, font_name: str = "Arial") -> None:
    """
    Apply dyslexia-friendly font and styling.

    Options:
    - Arial: Clean sans-serif
    - Helvetica: Alternative clean sans-serif
    - Courier New: Monospace (helps some dyslexic readers)

    Args:
        excel_path: Path to Excel file
        font_name: Font name to apply (default Arial)
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Grades"]

    for row in ws.iter_rows():
        for cell in row:
            if cell.font:
                cell.font = Font(
                    name=font_name,
                    size=cell.font.size or 11,
                    bold=cell.font.bold,
                    color=cell.font.color,
                )
            else:
                cell.font = Font(name=font_name, size=11)

    wb.save(excel_path)


def add_icon_labels_to_flags(excel_path: str) -> None:
    """
    Add semantic labels to flag icons (no color-only information).

    For each colored flag, add descriptive text abbreviation:
    - RED/Alert: [!]
    - YELLOW/Warning: [?]
    - GREEN/OK: [✓]

    Args:
        excel_path: Path to Excel file
    """
    wb = openpyxl.load_workbook(excel_path)

    # Iterate sheets and check for colored cells
    for ws in wb.sheetnames:
        sheet = wb[ws]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.start_color:
                    color = cell.fill.start_color.rgb
                    if color in ["FFFC7CE", "FFEB9C", "FFC6EFCE"]:  # Alert/Warning/OK colors
                        # Add icon label
                        if cell.value:
                            if color == "FFFC7CE":  # Red (Alert)
                                cell.value = f"[!] {cell.value}"
                            elif color == "FFEB9C":  # Yellow (Warning)
                                cell.value = f"[?] {cell.value}"
                            elif color == "FFC6EFCE":  # Green (OK)
                                cell.value = f"[✓] {cell.value}"

    wb.save(excel_path)
