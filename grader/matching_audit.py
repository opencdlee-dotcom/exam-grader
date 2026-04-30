"""
matching_audit.py — Generalized Matching Section Audit

Exam-agnostic module for auditing matching section scores.
Works with any exam, any section, any number of matching questions.

Matching Log Format (universal):
    "Q9=L✓ Q10=M✓ Q11=J✗(P) Q12=K✓ → 3/4"
    - ✓  = correct
    - ✗(X) = wrong; X is the correct letter
    - → N/total = score from log

Usage:
    from grader.matching_audit import run_matching_audit, build_matching_audit_sheet

    key = {"Q9": "L", "Q10": "M", "Q11": "P", ...}
    results = run_matching_audit(students, key)
    print(format_audit_report(results))
"""

import re
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Parsing ────────────────────────────────────────────────────────────────────

def parse_matching_log(log: str) -> dict:
    """
    Parse a matching log string into a dict of question → letter written.

    Input:  "Q9=L✓ Q10=M✓ Q11=J✗(P) Q12=K✓ Q13=O✗(H) → 3/5"
    Output: {"Q9": "L", "Q10": "M", "Q11": "J", "Q12": "K", "Q13": "O"}
    """
    result = {}
    # Match patterns like Q9=L, Q10=AB, Q11=J (letter before ✓ or ✗)
    pattern = re.compile(r'Q(\d+)=([A-Za-z]+)[✓✗]')
    for m in pattern.finditer(log):
        q_num = f"Q{m.group(1)}"
        letter = m.group(2).upper()
        result[q_num] = letter
    return result


def extract_matching_log_from_fr_details(fr_details: str) -> Optional[str]:
    """
    Extract the MATCHING: line from fr_details if present.

    Looks for lines starting with 'MATCHING:' and returns that segment.
    Returns None if no matching log is embedded.
    """
    for line in fr_details.splitlines():
        stripped = line.strip()
        if stripped.upper().startswith("MATCHING:"):
            return stripped[len("MATCHING:"):].strip()
    return None


# ── Scoring ────────────────────────────────────────────────────────────────────

def score_from_log(log: str, key: dict) -> int:
    """
    Compute correct count by parsing the log and comparing to key.

    Args:
        log: matching log string
        key: dict of {"Q9": "L", "Q10": "M", ...}
    Returns:
        integer count of correct answers (0 to len(key))
    """
    written = parse_matching_log(log)
    correct = 0
    for q, expected in key.items():
        if written.get(q, "").upper() == expected.upper():
            correct += 1
    return correct


def build_matching_log(written: dict, key: dict) -> str:
    """
    Build a formatted matching log string from a dict of written letters and the key.

    Args:
        written: {"Q9": "L", "Q10": "M", "Q11": "J", ...}  — what student wrote
        key:     {"Q9": "L", "Q10": "M", "Q11": "P", ...}  — correct answers
    Returns:
        "Q9=L✓ Q10=M✓ Q11=J✗(P) ... → 2/3"
    """
    parts = []
    correct = 0
    for q in sorted(key.keys(), key=lambda x: int(x[1:])):
        student_letter = written.get(q, "?").upper()
        expected = key[q].upper()
        if student_letter == expected:
            parts.append(f"{q}={student_letter}✓")
            correct += 1
        else:
            parts.append(f"{q}={student_letter}✗({expected})")
    total = len(key)
    return " ".join(parts) + f" → {correct}/{total}"


# ── Verification ───────────────────────────────────────────────────────────────

def verify_matching_score(student: dict, key: dict) -> dict:
    """
    Cross-check a student's recorded matching score against their matching log.

    Returns:
        {
            "name": str,
            "recorded": int | None,   # score stored in student dict (None if unknown)
            "computed": int | None,   # score from matching log (None if no log)
            "delta": int | None,      # computed - recorded (None if either missing)
            "status": "OK" | "DISCREPANCY" | "MISSING_LOG" | "NO_RECORDED_SCORE",
            "log": str | None,        # the matching log found
            "written": dict,          # parsed letters from log
        }
    """
    name = student.get("name", "Unknown")
    fr = student.get("fr_details", "")

    # Try to extract matching log
    log = extract_matching_log_from_fr_details(fr)
    written = parse_matching_log(log) if log else {}

    # Get recorded matching score — try matching_score field first, fall back to parsing
    recorded = student.get("matching_score")  # explicit field if present
    if recorded is None:
        # Try to parse from fr_details "Matching N/10" pattern
        m = re.search(r'[Mm]atching\s+(\d+)/\d+', fr)
        if m:
            recorded = int(m.group(1))

    computed = score_from_log(log, key) if log else None

    if log is None:
        status = "MISSING_LOG"
        delta = None
    elif recorded is None:
        status = "NO_RECORDED_SCORE"
        delta = None
    elif computed != recorded:
        status = "DISCREPANCY"
        delta = computed - recorded
    else:
        status = "OK"
        delta = 0

    return {
        "name": name,
        "recorded": recorded,
        "computed": computed,
        "delta": delta,
        "status": status,
        "log": log,
        "written": written,
    }


# ── Batch Audit ────────────────────────────────────────────────────────────────

def run_matching_audit(students: list, key: dict) -> list:
    """
    Audit matching scores for a list of students against a key.

    Args:
        students: list of student dicts (from generate_grades.py or JSON)
        key:      matching answer key, e.g. {"Q9": "L", "Q10": "M", ...}
    Returns:
        list of audit result dicts (one per student), sorted by |delta| descending
    """
    results = []
    for s in students:
        result = verify_matching_score(s, key)
        result["num"] = s.get("num", "?")
        results.append(result)

    # Sort: DISCREPANCY first (largest delta), then MISSING_LOG, then OK
    priority = {"DISCREPANCY": 0, "MISSING_LOG": 1, "NO_RECORDED_SCORE": 2, "OK": 3}
    results.sort(key=lambda r: (priority.get(r["status"], 9), -(abs(r["delta"]) if r["delta"] else 0)))
    return results


def format_audit_report(results: list, key: dict) -> str:
    """
    Return a human-readable audit report table.
    """
    lines = ["MATCHING AUDIT REPORT", "=" * 60]

    discrepancies = [r for r in results if r["status"] == "DISCREPANCY"]
    missing = [r for r in results if r["status"] == "MISSING_LOG"]
    ok = [r for r in results if r["status"] == "OK"]

    if discrepancies:
        lines.append(f"\n⚠ DISCREPANCIES ({len(discrepancies)} students):")
        for r in discrepancies:
            sign = "+" if r["delta"] > 0 else ""
            lines.append(f"  #{r['num']:>2} {r['name']:<30} recorded={r['recorded']} computed={r['computed']} Δ={sign}{r['delta']}")
            if r["log"]:
                lines.append(f"       {r['log']}")

    if missing:
        lines.append(f"\n? MISSING LOGS ({len(missing)} students — re-read required):")
        for r in missing:
            lines.append(f"  #{r['num']:>2} {r['name']:<30} recorded={r['recorded']}")

    lines.append(f"\n✓ OK ({len(ok)} students verified correct)")
    lines.append("=" * 60)
    return "\n".join(lines)


# ── Excel Sheet Builder ────────────────────────────────────────────────────────

def build_matching_audit_sheet(ws, students: list, key: dict, section: str):
    """
    Build a 'Matching Audit' Excel worksheet.

    Columns: #, Name, Q9, Q10, ..., Q18, Logged Score, Recorded, Δ, Status
    Key row shown at top with correct answers highlighted.
    Color coding: green=correct, red=wrong, yellow=missing, orange=discrepancy fixed.

    Args:
        ws:       openpyxl worksheet (already created, tab name set by caller)
        students: list of student dicts with matching_log in fr_details
        key:      {"Q9": "L", "Q10": "M", ...}
        section:  section name for title row
    """
    # Styles
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F3864")
    key_fill = PatternFill("solid", fgColor="2F75B6")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    orange_fill = PatternFill("solid", fgColor="FF9900")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="DCE6F1")

    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font = Font(name="Arial", size=10)
    bold_font = Font(bold=True, name="Arial", size=10)

    # Sort key questions numerically
    q_nums = sorted(key.keys(), key=lambda x: int(x[1:]))
    n_q = len(q_nums)

    # Column layout: #, Name, Q9..Q18, Logged, Recorded, Δ, Status
    col_num = 1
    col_name = 2
    col_q_start = 3
    col_q_end = col_q_start + n_q - 1
    col_logged = col_q_end + 1
    col_recorded = col_q_end + 2
    col_delta = col_q_end + 3
    col_status = col_q_end + 4
    total_cols = col_status

    # ── Title row ──
    ws.cell(1, 1).value = f"MATCHING AUDIT — {section}"
    ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="C00000")
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.row_dimensions[1].height = 24

    # ── Header row ──
    headers = ["#", "Name"] + q_nums + ["Logged", "Recorded", "Δ", "Status"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(2, col)
        cell.value = h
        cell.fill = header_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[2].height = 22

    # ── Key row ──
    ws.cell(3, col_num).value = "KEY"
    ws.cell(3, col_name).value = "(correct answers)"
    for i, q in enumerate(q_nums):
        cell = ws.cell(3, col_q_start + i)
        cell.value = key[q]
        cell.fill = key_fill
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for col in [col_num, col_name, col_logged, col_recorded, col_delta, col_status]:
        cell = ws.cell(3, col)
        cell.fill = key_fill
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[3].height = 20

    # ── Student rows ──
    audit_results = run_matching_audit(students, key)
    result_by_name = {r["name"]: r for r in audit_results}

    for i, s in enumerate(students):
        row = 4 + i
        alt = (i % 2 == 1)
        base_fill = alt_fill if alt else white_fill

        result = result_by_name.get(s.get("name", ""), {})
        written = result.get("written", {})
        status = result.get("status", "MISSING_LOG")
        delta = result.get("delta")
        logged = result.get("computed")
        recorded = result.get("recorded")

        # # and Name
        ws.cell(row, col_num).value = s.get("num", i + 1)
        ws.cell(row, col_name).value = s.get("name", "")

        for col in [col_num, col_name]:
            c = ws.cell(row, col)
            c.fill = base_fill
            c.font = body_font
            c.alignment = Alignment(vertical="center")
            c.border = border

        # Per-question columns
        for j, q in enumerate(q_nums):
            col = col_q_start + j
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            cell.font = bold_font

            letter = written.get(q)
            if letter is None:
                cell.value = "?"
                cell.fill = yellow_fill
            elif letter.upper() == key[q].upper():
                cell.value = letter
                cell.fill = green_fill
            else:
                cell.value = letter
                cell.fill = red_fill

        # Logged, Recorded, Δ, Status
        ws.cell(row, col_logged).value = logged if logged is not None else "—"
        ws.cell(row, col_recorded).value = recorded if recorded is not None else "—"

        delta_cell = ws.cell(row, col_delta)
        if delta is not None and delta != 0:
            delta_cell.value = f"+{delta}" if delta > 0 else str(delta)
            delta_cell.fill = orange_fill
            delta_cell.font = Font(bold=True, name="Arial", size=10)
        else:
            delta_cell.value = delta if delta is not None else "?"
            delta_cell.fill = base_fill
            delta_cell.font = body_font

        status_cell = ws.cell(row, col_status)
        status_cell.value = status
        if status == "OK":
            status_cell.fill = green_fill
        elif status == "DISCREPANCY":
            status_cell.fill = orange_fill
            status_cell.font = Font(bold=True, name="Arial", size=10)
        elif status == "MISSING_LOG":
            status_cell.fill = yellow_fill
        else:
            status_cell.fill = base_fill

        for col in [col_logged, col_recorded, col_delta, col_status]:
            c = ws.cell(row, col)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            if c.font is None or c.font == body_font:
                c.font = body_font

        ws.row_dimensions[row].height = 18

    # ── Column widths ──
    ws.column_dimensions[get_column_letter(col_num)].width = 5
    ws.column_dimensions[get_column_letter(col_name)].width = 28
    for j in range(n_q):
        ws.column_dimensions[get_column_letter(col_q_start + j)].width = 7
    ws.column_dimensions[get_column_letter(col_logged)].width = 9
    ws.column_dimensions[get_column_letter(col_recorded)].width = 11
    ws.column_dimensions[get_column_letter(col_delta)].width = 6
    ws.column_dimensions[get_column_letter(col_status)].width = 18

    ws.freeze_panes = "C4"


# ── Convenience: key builders ──────────────────────────────────────────────────

def make_key(q_start: int, letters: list) -> dict:
    """
    Build a key dict from a starting question number and list of correct letters.

    Example:
        make_key(9, ["L","M","P","K","O","H","G","B","D","I"])
        → {"Q9": "L", "Q10": "M", ..., "Q18": "I"}
    """
    return {f"Q{q_start + i}": letter for i, letter in enumerate(letters)}
