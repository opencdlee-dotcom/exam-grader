"""
score_breakdown_audit.py — Generalized Full Score Breakdown Audit

Exam-agnostic module for auditing all sections of any exam.
Works with any exam, any section layout, any point structure.

Sections are described by an ExamSchema (list of SectionSchema objects).
The same code works for a 35-pt bio exam, a 50-pt chem exam, or any future format.

Log Format (universal, one line per section in fr_details):
    MATCHING: Q9=L✓ Q10=M✓ Q11=J✗(P) → 9/10
    FILL-IN: Q19=denaturation Q20=annealing Q21=extension Q22=replication Q23=polymerase → 4/5
    Q24: ATCGGCATCG → 2/2
    Q25: Leucine/Tyrosine/Stop → 3/3
    Q26: agarose gel electrophoresis → 2/2
    FR: [1,0,1,0,1] → 3/5
    TOTAL: MC(8) + MATCHING(9) + FILL-IN(4) + Q24(2) + Q25(3) + Q26(2) + FR(3) = 31/35

Usage:
    from grader.score_breakdown_audit import (
        ExamSchema, SectionSchema,
        run_full_audit, assign_confidence,
        build_score_breakdown_sheet,
        pre_post_confidence_gate,
    )

    schema = ExamSchema(
        exam_name="Exam 2", section="T5", total_points=35,
        sections=[
            SectionSchema("MC",       "mc",       [f"Q{i}" for i in range(1,9)],
                          {"Q1":"B","Q2":"A",...}, 1.0, 8),
            SectionSchema("MATCHING", "matching", [f"Q{i}" for i in range(9,19)],
                          {"Q9":"L",...}, 1.0, 10),
            ...
        ]
    )
    results = run_full_audit(students, schema)
"""

import re
from dataclasses import dataclass, field
from typing import Optional, Union
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Schema Definitions ─────────────────────────────────────────────────────────

@dataclass
class SectionSchema:
    """
    Describes one scoreable section of an exam.

    Args:
        name:           Display name, e.g. "MC", "MATCHING", "FILL-IN", "FR", "Q24"
        section_type:   One of: "mc" | "matching" | "fill_in" | "fr" | "numeric"
        questions:      List of question labels, e.g. ["Q9","Q10",...,"Q18"]
        key:            Answer key.
                          - mc/matching/fill_in: dict of {question: correct_answer_str}
                          - fr: list of int (1=point possible, 0=not used, len=num bullets)
                            OR int (just the max points, no per-bullet key available)
                          - numeric: dict of {question: accepted_answer_or_list}
        points_per_q:   Points awarded per correct question (default 1.0)
        max_points:     Maximum points this section is worth
    """
    name: str
    section_type: str
    questions: list
    key: Union[dict, list, int]
    points_per_q: float = 1.0
    max_points: float = 0.0

    def __post_init__(self):
        if self.max_points == 0.0:
            self.max_points = len(self.questions) * self.points_per_q


@dataclass
class ExamSchema:
    """
    Full exam structure: a list of sections plus metadata.

    Args:
        exam_name:    e.g. "Exam 2"
        section:      e.g. "T5", "W2", "W5", "Section A"
        total_points: Total possible points across all sections
        sections:     Ordered list of SectionSchema objects
    """
    exam_name: str
    section: str
    total_points: float
    sections: list = field(default_factory=list)

    def get_section(self, name: str) -> Optional[SectionSchema]:
        """Return the SectionSchema with the given name, or None."""
        for s in self.sections:
            if s.name.upper() == name.upper():
                return s
        return None

    def section_names(self) -> list:
        return [s.name for s in self.sections]


# ── Log Parsing ────────────────────────────────────────────────────────────────

def parse_section_log(log_line: str, schema: SectionSchema) -> dict:
    """
    Parse one log line from fr_details for the given section type.

    Returns a dict of {question_label: written_value_str}.
    For FR sections, returns {"_bullets": [0, 1, 0, ...], "_sum": N}.
    For numeric sections, returns {question_label: written_str}.
    Returns {} if the line cannot be parsed.

    Examples:
        "MATCHING: Q9=L✓ Q10=M✗(P) → 9/10"  → {"Q9": "L", "Q10": "M", ...}
        "FILL-IN: Q19=denaturation Q20=annealing → 2/5"  → {"Q19": "denaturation", ...}
        "FR: [1,0,1,0,1] → 3/5"  → {"_bullets": [1,0,1,0,1], "_sum": 3}
        "Q24: ATCGGCATCG → 2/2"  → {"Q24": "ATCGGCATCG"}
    """
    if not log_line:
        return {}

    # Strip the section prefix (e.g. "MATCHING: " or "Q24: ")
    colon_idx = log_line.find(":")
    if colon_idx == -1:
        return {}
    body = log_line[colon_idx + 1:].strip()

    # Strip trailing "→ N/max" score
    body = re.sub(r'→\s*[\d.]+/[\d.]+\s*$', '', body).strip()
    body = re.sub(r'=\s*[\d.]+/[\d.]+\s*$', '', body).strip()

    result = {}

    if schema.section_type in ("mc", "matching", "fill_in"):
        # Pattern: Qn=value (optionally followed by ✓ or ✗(...))
        pattern = re.compile(r'(Q\d+)=([^\s✓✗]+)[✓✗]?(?:\([^)]*\))?')
        for m in pattern.finditer(body):
            q = m.group(1)
            val = m.group(2).strip()
            result[q] = val

    elif schema.section_type == "fr":
        # Pattern: [1,0,1,0,1] or [pt1=1,pt2=0,...]
        m = re.search(r'\[([^\]]+)\]', body)
        if m:
            raw = m.group(1)
            # Handle both "[1,0,1]" and "[pt1=1,pt2=0,pt3=1]"
            nums = re.findall(r'(?:pt\d+=)?(\d)', raw)
            bullets = [int(n) for n in nums]
            result["_bullets"] = bullets
            result["_sum"] = sum(bullets)
        else:
            # Fallback: just a number written (no bullet breakdown)
            m2 = re.search(r'(\d+)', body)
            if m2:
                result["_sum"] = int(m2.group(1))

    elif schema.section_type == "numeric":
        # The body after stripping is the student's written answer
        if schema.questions:
            q_label = schema.questions[0] if len(schema.questions) == 1 else schema.name
            result[q_label] = body.strip()

    return result


def extract_section_log_line(fr_details: str, section_name: str) -> Optional[str]:
    """
    Extract the log line for a named section from fr_details.

    Searches for lines starting with SECTION_NAME: (case-insensitive).
    Returns the full line, or None if not found.
    """
    for line in fr_details.splitlines():
        stripped = line.strip()
        if stripped.upper().startswith(section_name.upper() + ":"):
            return stripped
    return None


def extract_total_line(fr_details: str) -> Optional[str]:
    """
    Extract the TOTAL line from fr_details.

    Expected format: "TOTAL: MC(8) + MATCHING(9) + ... = 31/35"
    Returns the full line, or None if not found.
    """
    for line in fr_details.splitlines():
        stripped = line.strip()
        if stripped.upper().startswith("TOTAL:"):
            return stripped
    return None


# ── Scoring ────────────────────────────────────────────────────────────────────

def score_section_log(parsed: dict, schema: SectionSchema) -> Optional[float]:
    """
    Compute earned points from a parsed section log vs the schema key.

    Returns:
        float: computed score for this section
        None:  cannot compute (empty parsed, no key)
    """
    if not parsed:
        return None

    if schema.section_type in ("mc", "matching", "fill_in"):
        if not isinstance(schema.key, dict):
            return None
        correct = 0
        for q, expected in schema.key.items():
            written = parsed.get(q, "").strip().upper()
            if written and written == expected.strip().upper():
                correct += 1
        return correct * schema.points_per_q

    elif schema.section_type == "fr":
        if "_sum" in parsed:
            return float(parsed["_sum"])
        return None

    elif schema.section_type == "numeric":
        # Numeric sections: check if written answer matches any accepted form
        if not schema.key:
            return None
        q_label = schema.questions[0] if len(schema.questions) == 1 else schema.name
        written = parsed.get(q_label, "").strip().lower()
        if not written:
            return None
        accepted = schema.key
        if isinstance(accepted, dict):
            accepted = accepted.get(q_label, accepted.get(schema.name, ""))
        if isinstance(accepted, list):
            accepted_normalized = [str(a).strip().lower() for a in accepted]
            return schema.max_points if written in accepted_normalized else 0.0
        else:
            return schema.max_points if written == str(accepted).strip().lower() else 0.0

    return None


def parse_total_line(total_line: str) -> dict:
    """
    Parse a TOTAL line into its section components.

    Input:  "TOTAL: MC(8) + MATCHING(9) + FILL-IN(4) + Q24(2) + FR(3) = 31/35"
    Output: {
        "sections": {"MC": 8, "MATCHING": 9, "FILL-IN": 4, "Q24": 2, "FR": 3},
        "reported_total": 31,
        "reported_max": 35,
    }
    """
    result = {"sections": {}, "reported_total": None, "reported_max": None}

    colon_idx = total_line.find(":")
    if colon_idx == -1:
        return result
    body = total_line[colon_idx + 1:].strip()

    # Extract section(score) pairs
    for m in re.finditer(r'([A-Z0-9\-]+)\((\d+(?:\.\d+)?)\)', body):
        result["sections"][m.group(1)] = float(m.group(2))

    # Extract "= N/M" at the end
    m = re.search(r'=\s*([\d.]+)/([\d.]+)\s*$', body)
    if m:
        result["reported_total"] = float(m.group(1))
        result["reported_max"] = float(m.group(2))

    return result


# ── Per-Student Audit ──────────────────────────────────────────────────────────

def verify_student_breakdown(student: dict, schema: ExamSchema) -> dict:
    """
    Full audit of one student against the exam schema.

    For each section:
      - Extracts the log line from fr_details
      - Parses it
      - Computes score vs key (where applicable)
      - Compares computed to what the TOTAL line reported for that section

    Also verifies the TOTAL line arithmetic.

    Returns:
        {
            "name": str,
            "num": str | int,
            "sections": {
                section_name: {
                    "log_found": bool,
                    "parsed": dict,
                    "computed": float | None,
                    "reported": float | None,  # from TOTAL line
                    "delta": float | None,
                    "status": "OK" | "DISCREPANCY" | "MISSING_LOG" | "UNVERIFIABLE",
                }
            },
            "total": {
                "reported": float | None,
                "computed": float | None,  # sum of section computed scores
                "delta": float | None,
                "status": "OK" | "DISCREPANCY" | "MISSING_TOTAL",
            },
            "confidence": "HIGH" | "MEDIUM" | "LOW",
            "worst_status": str,
        }
    """
    name = student.get("name", "Unknown")
    num = student.get("num", "?")
    fr_details = student.get("fr_details", "")

    # Parse total line
    total_line = extract_total_line(fr_details)
    total_parsed = parse_total_line(total_line) if total_line else {}
    total_section_scores = total_parsed.get("sections", {})

    sections_result = {}
    computed_section_scores = []

    for sec in schema.sections:
        log_line = extract_section_log_line(fr_details, sec.name)
        log_found = log_line is not None

        parsed = parse_section_log(log_line, sec) if log_found else {}
        computed = score_section_log(parsed, sec) if log_found else None

        # Reported score for this section from the TOTAL line
        reported = total_section_scores.get(sec.name.upper())

        if computed is not None:
            computed_section_scores.append(computed)

        # Determine status
        if not log_found:
            status = "MISSING_LOG"
            delta = None
        elif computed is None:
            status = "UNVERIFIABLE"
            delta = None
        elif reported is None:
            # Log found but no TOTAL line breakdown — can't cross-check
            status = "OK"  # trust the log, can't verify arithmetic
            delta = None
        elif abs(computed - reported) < 0.01:
            status = "OK"
            delta = 0.0
        else:
            status = "DISCREPANCY"
            delta = computed - reported

        sections_result[sec.name] = {
            "log_found": log_found,
            "parsed": parsed,
            "computed": computed,
            "reported": reported,
            "delta": delta,
            "status": status,
        }

    # Verify TOTAL line arithmetic
    total_computed = sum(computed_section_scores) if computed_section_scores else None
    total_reported = total_parsed.get("reported_total")

    if total_line is None:
        total_status = "MISSING_TOTAL"
        total_delta = None
    elif total_reported is None:
        total_status = "MISSING_TOTAL"
        total_delta = None
    elif total_computed is None:
        total_status = "UNVERIFIABLE"
        total_delta = None
    elif abs(total_computed - total_reported) < 0.01:
        total_status = "OK"
        total_delta = 0.0
    else:
        total_status = "DISCREPANCY"
        total_delta = total_computed - total_reported

    total_result = {
        "reported": total_reported,
        "computed": total_computed,
        "delta": total_delta,
        "status": total_status,
    }

    # Determine overall worst status
    all_statuses = [s["status"] for s in sections_result.values()] + [total_status]
    priority = {"DISCREPANCY": 0, "MISSING_LOG": 1, "MISSING_TOTAL": 2, "UNVERIFIABLE": 3, "OK": 4}
    worst = min(all_statuses, key=lambda s: priority.get(s, 9))

    # Assign confidence
    confidence = assign_confidence_from_result(sections_result, total_result, schema)

    return {
        "name": name,
        "num": num,
        "sections": sections_result,
        "total": total_result,
        "confidence": confidence,
        "worst_status": worst,
    }


def assign_confidence_from_result(sections_result: dict, total_result: dict, schema: ExamSchema) -> str:
    """
    HIGH: all sections have logs + TOTAL line present + no DISCREPANCY
    MEDIUM: ≥60% of sections have logs, no hard discrepancies
    LOW: <40% of sections have logs OR any DISCREPANCY
    """
    n_sections = len(schema.sections)
    if n_sections == 0:
        return "LOW"

    logs_found = sum(1 for s in sections_result.values() if s.get("log_found"))
    has_discrepancy = any(
        s["status"] == "DISCREPANCY" for s in sections_result.values()
    ) or total_result.get("status") == "DISCREPANCY"

    pct = logs_found / n_sections
    has_total = total_result.get("status") not in ("MISSING_TOTAL",)

    if has_discrepancy:
        return "LOW"
    if pct >= 1.0 and has_total:
        return "HIGH"
    if pct >= 0.6:
        return "MEDIUM"
    return "LOW"


def assign_confidence(student: dict, schema: ExamSchema) -> str:
    """
    Convenience wrapper: run audit on one student and return its confidence level.
    """
    result = verify_student_breakdown(student, schema)
    return result["confidence"]


# ── Batch Audit ────────────────────────────────────────────────────────────────

def run_full_audit(students: list, schema: ExamSchema) -> list:
    """
    Run full score breakdown audit on a list of students.

    Returns list of audit result dicts (one per student), sorted:
        DISCREPANCY first, then MISSING_LOG, then MISSING_TOTAL, then OK.
    """
    results = []
    for s in students:
        result = verify_student_breakdown(s, schema)
        results.append(result)

    priority = {"DISCREPANCY": 0, "MISSING_LOG": 1, "MISSING_TOTAL": 2, "UNVERIFIABLE": 3, "OK": 4}
    results.sort(key=lambda r: priority.get(r["worst_status"], 9))
    return results


def format_full_audit_report(results: list, schema: ExamSchema) -> str:
    """Return a human-readable full breakdown audit report."""
    lines = [f"SCORE BREAKDOWN AUDIT — {schema.exam_name} {schema.section}", "=" * 70]

    discrepancies = [r for r in results if r["worst_status"] == "DISCREPANCY"]
    missing = [r for r in results if r["worst_status"] in ("MISSING_LOG", "MISSING_TOTAL")]
    ok = [r for r in results if r["worst_status"] == "OK"]

    if discrepancies:
        lines.append(f"\n⚠ DISCREPANCIES ({len(discrepancies)} students):")
        for r in discrepancies:
            lines.append(f"  #{r['num']:>2} {r['name']:<30}")
            for sec_name, sec in r["sections"].items():
                if sec["status"] == "DISCREPANCY":
                    sign = "+" if sec["delta"] > 0 else ""
                    lines.append(
                        f"       {sec_name}: computed={sec['computed']} "
                        f"reported={sec['reported']} Δ={sign}{sec['delta']}"
                    )
            if r["total"]["status"] == "DISCREPANCY":
                sign = "+" if r["total"]["delta"] > 0 else ""
                lines.append(
                    f"       TOTAL: computed={r['total']['computed']} "
                    f"reported={r['total']['reported']} Δ={sign}{r['total']['delta']}"
                )

    if missing:
        lines.append(f"\n? MISSING LOGS ({len(missing)} students):")
        for r in missing:
            missing_secs = [n for n, s in r["sections"].items() if not s["log_found"]]
            lines.append(
                f"  #{r['num']:>2} {r['name']:<30} missing: {', '.join(missing_secs)}"
                + (" [NO TOTAL LINE]" if r["total"]["status"] == "MISSING_TOTAL" else "")
            )

    lines.append(f"\n✓ OK — {len(ok)} students fully verified")
    lines.append("=" * 70)
    return "\n".join(lines)


# ── Pre-Post Gate ──────────────────────────────────────────────────────────────

def pre_post_confidence_gate(students: list, schema: ExamSchema) -> bool:
    """
    Check confidence levels before Canvas posting.

    Returns True if safe to proceed, False if user cancelled.
    Prints a warning table for LOW-confidence or DISCREPANCY-flagged students.
    Requires explicit 'yes' to proceed if any LOW-confidence students found.

    This is a soft gate — the instructor can override — but requires
    explicit acknowledgment of every LOW-confidence student.
    """
    low = [s for s in students if s.get("confidence", "LOW") == "LOW"]
    discrepant = [
        s for s in students
        if "MATCHING_AUDIT" in s.get("flags", "") or "SCORE_AUDIT" in s.get("flags", "")
    ]

    if not low and not discrepant:
        print(f"✓ Pre-post gate: all {len(students)} students are MEDIUM or HIGH confidence. Safe to post.")
        return True

    print(f"\n{'='*60}")
    print("⚠  PRE-POST CONFIDENCE GATE")
    print(f"{'='*60}")

    if low:
        print(f"\n  {len(low)} LOW-confidence students (unverified grades):")
        for s in low:
            score = s.get("score", "?")
            pct = s.get("pct", "?")
            pct_str = f"{pct:.1f}%" if isinstance(pct, (int, float)) else str(pct)
            print(f"    #{s.get('num','?'):>2} {s.get('name','Unknown'):<30} "
                  f"{score}/{schema.total_points} ({pct_str})")

    if discrepant:
        print(f"\n  {len(discrepant)} students with unresolved audit flags:")
        for s in discrepant:
            print(f"    #{s.get('num','?'):>2} {s.get('name','Unknown'):<30} "
                  f"flags: {s.get('flags','')}")

    print(f"\n  Posting unverified grades may publish incorrect scores to Canvas.")
    print(f"  Resolve LOW-confidence students first, or override with 'yes'.\n")

    try:
        confirm = input("Post grades anyway? (yes/no): ").strip().lower()
    except (EOFError, KeyboardInterrupt):
        print("\nPosting cancelled.")
        return False

    if confirm == "yes":
        print(f"Override confirmed. Posting {len(students)} grades.")
        return True
    else:
        print("Posting cancelled. Resolve LOW-confidence students and retry.")
        return False


# ── Excel Sheet Builder ────────────────────────────────────────────────────────

def build_score_breakdown_sheet(ws, students: list, schema: ExamSchema):
    """
    Build a 'Score Breakdown' Excel worksheet.

    Columns: #, Name, [one col per section], TOTAL, Confidence, Status
    Color coding:
        - Confidence: green=HIGH, yellow=MEDIUM, red=LOW
        - Sections: green=OK, orange=DISCREPANCY, yellow=MISSING_LOG, white=UNVERIFIABLE
        - TOTAL: green=OK, orange=DISCREPANCY, yellow=MISSING_TOTAL
    """
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F3864")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    orange_fill = PatternFill("solid", fgColor="FF9900")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="DCE6F1")

    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font = Font(name="Arial", size=10)
    bold_font = Font(bold=True, name="Arial", size=10)

    section_names = schema.section_names()
    n_sections = len(section_names)

    # Column layout
    col_num = 1
    col_name = 2
    col_sec_start = 3
    col_sec_end = col_sec_start + n_sections - 1
    col_total = col_sec_end + 1
    col_conf = col_sec_end + 2
    col_status = col_sec_end + 3
    total_cols = col_status

    # ── Title row ──
    ws.cell(1, 1).value = f"SCORE BREAKDOWN — {schema.exam_name} {schema.section}"
    ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", name="Arial", size=13)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="1F5C1A")
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.row_dimensions[1].height = 24

    # ── Header row ──
    headers = ["#", "Name"] + section_names + ["TOTAL", "Confidence", "Status"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(2, col)
        cell.value = h
        cell.fill = header_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[2].height = 22

    # ── Max pts row ──
    ws.cell(3, col_num).value = "MAX"
    ws.cell(3, col_name).value = "(max points per section)"
    for i, sec in enumerate(schema.sections):
        cell = ws.cell(3, col_sec_start + i)
        cell.value = sec.max_points
        cell.fill = PatternFill("solid", fgColor="2F75B6")
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.cell(3, col_total).value = schema.total_points
    ws.cell(3, col_total).fill = PatternFill("solid", fgColor="2F75B6")
    ws.cell(3, col_total).font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    ws.cell(3, col_total).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(3, col_total).border = border
    for col in [col_num, col_name, col_conf, col_status]:
        c = ws.cell(3, col)
        c.fill = PatternFill("solid", fgColor="2F75B6")
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws.row_dimensions[3].height = 20

    # ── Student rows ──
    audit_results = run_full_audit(students, schema)
    result_by_name = {r["name"]: r for r in audit_results}

    for i, s in enumerate(students):
        row = 4 + i
        alt = (i % 2 == 1)
        base_fill = alt_fill if alt else white_fill

        result = result_by_name.get(s.get("name", ""), {})
        sections_data = result.get("sections", {})
        total_data = result.get("total", {})
        confidence = result.get("confidence", "LOW")
        worst = result.get("worst_status", "MISSING_LOG")

        # # and Name
        ws.cell(row, col_num).value = s.get("num", i + 1)
        ws.cell(row, col_name).value = s.get("name", "")
        for col in [col_num, col_name]:
            c = ws.cell(row, col)
            c.fill = base_fill
            c.font = body_font
            c.alignment = Alignment(vertical="center")
            c.border = border

        # Section columns
        for j, sec_name in enumerate(section_names):
            col = col_sec_start + j
            cell = ws.cell(row, col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

            sec_data = sections_data.get(sec_name, {})
            computed = sec_data.get("computed")
            status = sec_data.get("status", "MISSING_LOG")
            delta = sec_data.get("delta")

            if not sec_data.get("log_found"):
                cell.value = "?"
                cell.fill = yellow_fill
                cell.font = body_font
            elif status == "DISCREPANCY":
                sign = "+" if delta > 0 else ""
                cell.value = f"{computed} (Δ{sign}{delta:.0f})"
                cell.fill = orange_fill
                cell.font = bold_font
            elif computed is not None:
                cell.value = computed
                sec_schema = schema.get_section(sec_name)
                max_pts = sec_schema.max_points if sec_schema else None
                if max_pts and computed >= max_pts:
                    cell.fill = green_fill
                else:
                    cell.fill = base_fill
                cell.font = body_font
            else:
                cell.value = "—"
                cell.fill = base_fill
                cell.font = body_font

        # TOTAL column
        total_cell = ws.cell(row, col_total)
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        total_cell.border = border
        t_reported = total_data.get("reported")
        t_status = total_data.get("status", "MISSING_TOTAL")
        t_delta = total_data.get("delta")

        if t_status == "MISSING_TOTAL":
            total_cell.value = s.get("score", "?")
            total_cell.fill = yellow_fill
            total_cell.font = body_font
        elif t_status == "DISCREPANCY":
            sign = "+" if t_delta > 0 else ""
            total_cell.value = f"{t_reported} (Δ{sign}{t_delta:.0f})"
            total_cell.fill = orange_fill
            total_cell.font = bold_font
        else:
            total_cell.value = t_reported if t_reported is not None else s.get("score", "?")
            total_cell.fill = green_fill if t_status == "OK" else base_fill
            total_cell.font = body_font

        # Confidence column
        conf_cell = ws.cell(row, col_conf)
        conf_cell.value = confidence
        conf_cell.alignment = Alignment(horizontal="center", vertical="center")
        conf_cell.border = border
        if confidence == "HIGH":
            conf_cell.fill = green_fill
            conf_cell.font = body_font
        elif confidence == "MEDIUM":
            conf_cell.fill = yellow_fill
            conf_cell.font = body_font
        else:
            conf_cell.fill = red_fill
            conf_cell.font = bold_font

        # Status column
        status_cell = ws.cell(row, col_status)
        status_cell.value = worst
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        status_cell.border = border
        if worst == "OK":
            status_cell.fill = green_fill
            status_cell.font = body_font
        elif worst == "DISCREPANCY":
            status_cell.fill = orange_fill
            status_cell.font = bold_font
        elif worst in ("MISSING_LOG", "MISSING_TOTAL"):
            status_cell.fill = yellow_fill
            status_cell.font = body_font
        else:
            status_cell.fill = base_fill
            status_cell.font = body_font

        ws.row_dimensions[row].height = 18

    # ── Column widths ──
    ws.column_dimensions[get_column_letter(col_num)].width = 5
    ws.column_dimensions[get_column_letter(col_name)].width = 28
    for j in range(n_sections):
        ws.column_dimensions[get_column_letter(col_sec_start + j)].width = 12
    ws.column_dimensions[get_column_letter(col_total)].width = 12
    ws.column_dimensions[get_column_letter(col_conf)].width = 13
    ws.column_dimensions[get_column_letter(col_status)].width = 16

    ws.freeze_panes = "C4"
